"""Tests for the small helpers in app/ingestion/parser.py."""
import io
from datetime import date

import openpyxl
import pytest

from app.ingestion.parser import (
    _clean_num,
    _clean_str,
    _normalize,
    _col_letter_to_index,
    _parse_shift_line,
    _dump_cells,
    _find_cell,
    _find_all_cells,
    parse_workbook,
)


class TestCleanNum:
    @pytest.mark.parametrize(
        "value, expected",
        [
            (1, 1.0),
            (1.5, 1.5),
            ("3.14", 3.14),
            ("  42 ", 42.0),
            ("0", 0.0),
            ("-5", -5.0),
        ],
    )
    def test_returns_float(self, value, expected):
        assert _clean_num(value) == expected

    @pytest.mark.parametrize(
        "value",
        [None, "", "-", "—", "#DIV/0!", "#N/A", "#REF!", "#VALUE!", "abc", "1.2.3"],
    )
    def test_returns_none(self, value):
        assert _clean_num(value) is None


class TestCleanStr:
    @pytest.mark.parametrize(
        "value, expected",
        [
            ("hello", "hello"),
            ("  spaced  ", "spaced"),
            ("MIX 12345", "MIX 12345"),
            (123, "123"),
        ],
    )
    def test_returns_string(self, value, expected):
        assert _clean_str(value) == expected

    @pytest.mark.parametrize(
        "value",
        [None, "", "  ", "-", "—", "#DIV/0!"],
    )
    def test_returns_none(self, value):
        assert _clean_str(value) is None


class TestNormalize:
    def test_collapses_internal_whitespace(self):
        assert _normalize("شیفت    روز") == "شیفت روز"

    def test_replaces_newline_with_space(self):
        assert _normalize("شیفت\nروز") == "شیفت روز"

    def test_strips_outer_whitespace(self):
        assert _normalize("  text  ") == "text"

    def test_handles_none(self):
        assert _normalize(None) == ""

    def test_handles_empty(self):
        assert _normalize("") == ""


class TestColLetterToIndex:
    @pytest.mark.parametrize(
        "letter, expected",
        [
            ("A", 1),
            ("B", 2),
            ("Z", 26),
            ("AA", 27),
            ("AB", 28),
            ("AC", 29),
            ("AD", 30),
            ("AZ", 52),
            ("BA", 53),
        ],
    )
    def test_letter_to_index(self, letter, expected):
        assert _col_letter_to_index(letter) == expected

    def test_lowercase_works(self):
        assert _col_letter_to_index("aa") == 27


class TestParseShiftLine:
    def test_day_shift_line_one(self):
        assert _parse_shift_line("شیفت روز", "خط 1") == ("day", 1)

    def test_night_shift_line_two(self):
        assert _parse_shift_line("شیفت شب", "خط 2") == ("night", 2)

    def test_total(self):
        assert _parse_shift_line("جمع کل", "") == ("total", None)

    def test_unknown_shift(self):
        assert _parse_shift_line("", "خط 1") == (None, 1)

    def test_no_line_number(self):
        assert _parse_shift_line("شیفت روز", "no digits here") == ("day", None)


class TestDumpCells:
    def test_only_non_empty_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["B2"] = 42
        ws["C3"] = ""  # blank string should be skipped
        ws["D4"] = "   "  # whitespace-only should be skipped
        ws["E5"] = 3.14

        cells = _dump_cells(ws)

        assert cells == {"A1": "hello", "B2": 42, "E5": 3.14}

    def test_dump_is_json_friendly(self):
        import json

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Persian: شیفت روز"
        ws["B1"] = 100
        ws["C1"] = True

        cells = _dump_cells(ws)
        # Must be JSON-serialisable (gets stored as JSONB)
        json.dumps(cells, ensure_ascii=False)


class TestFindCell:
    def test_finds_anchor_with_exact_match(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B5"] = "تاریخ گزارش"

        cell = _find_cell(ws, "تاریخ")
        assert cell is not None
        assert cell.coordinate == "B5"

    def test_returns_none_when_missing(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "something else"
        assert _find_cell(ws, "تاریخ") is None

    def test_match_is_whitespace_tolerant(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "علت    توقف   فید"
        cell = _find_cell(ws, "علت توقف فید")
        assert cell is not None


class TestFindAllCells:
    def test_finds_all_matching_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "دلایل توقف فیلتر پرس"
        ws["Q1"] = "دلایل توقف فیلتر پرس - شب"

        found = _find_all_cells(ws, "دلایل توقف فیلتر پرس")
        coords = sorted(c.coordinate for c in found)
        assert coords == ["A1", "Q1"]


class TestParseWorkbookEndToEnd:
    """Build a minimal synthetic sheet that exercises the full parser."""

    def _build_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "01_05"

        # Header: date anchor in column A, date in column B
        ws.cell(row=1, column=1, value="تاریخ")
        ws.cell(row=1, column=2, value="1405/01/05")

        # Production block: shift label A7-A8, line in B
        ws.cell(row=7, column=1, value="شیفت روز")
        ws.cell(row=7, column=2, value="خط 1")
        ws.cell(row=7, column=3, value=1000)   # daily_feed_tonnage
        ws.cell(row=7, column=5, value=750)    # daily_concentrate_tonnage
        ws.cell(row=7, column=7, value=0.75)   # daily_recovery_percent

        ws.cell(row=8, column=1, value="شیفت روز")
        ws.cell(row=8, column=2, value="خط 2")
        ws.cell(row=8, column=3, value=900)
        ws.cell(row=8, column=5, value=680)
        ws.cell(row=8, column=7, value=0.755)

        # Quality block — anchor at row 50
        ws.cell(row=49, column=1, value="کیفیت خوراک")
        ws.cell(row=51, column=1, value="شیفت روز")
        ws.cell(row=51, column=2, value="خط 1")
        ws.cell(row=51, column=3, value=35.5)  # feed_fe_percent
        ws.cell(row=51, column=5, value=66.99)  # concentrate_fe_percent

        # Factory downtime: anchor at L13 (col 12)
        ws.cell(row=13, column=12, value="دلایل توقف کارخانه")
        ws.cell(row=14, column=12, value="شیفت روز")
        ws.cell(row=14, column=13, value="خط 1")
        ws.cell(row=14, column=14, value="خرابی برق پمپ شماره 5 از ساعت 07:00 تا 09:00")  # text in N
        ws.cell(row=14, column=23, value=120)  # duration in W (col 12 + 11 = 23)

        # Batch code in last row
        ws.cell(row=58, column=1, value="MIX (MAHCOARSE030711_MAH040701225")

        return wb

    def test_full_parse(self, tmp_path):
        wb = self._build_sheet()
        path = tmp_path / "test_factory.xlsx"
        wb.save(path)

        result = parse_workbook(path)

        assert result.sheets_parsed == 1
        assert result.sheets_failed == 0
        assert len(result.daily_reports) == 1

        header = result.daily_reports[0]
        assert header["report_date"] == date(2026, 3, 25)
        assert header["jalali_date"] == "1405/01/05"
        assert header["sheet_name"] == "01_05"
        assert header["batch_code"] is not None
        assert "MIX" in header["batch_code"]

    def test_production_rows_extracted(self, tmp_path):
        wb = self._build_sheet()
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = parse_workbook(path)

        # Two production rows for two lines (plus quality merged in)
        assert len(result.production_rows) >= 2
        line1 = next(r for r in result.production_rows if r.get("line") == 1 and r.get("shift") == "day")
        assert line1["daily_feed_tonnage"] == 1000
        assert line1["daily_concentrate_tonnage"] == 750

    def test_downtime_row_extracted(self, tmp_path):
        wb = self._build_sheet()
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = parse_workbook(path)

        factory_events = [d for d in result.downtime_rows if d["section"] == "factory"]
        assert len(factory_events) >= 1
        event = factory_events[0]
        assert event["duration_minutes"] == 120
        assert event["start_time"] == "07:00"
        assert event["end_time"] == "09:00"
        assert "خرابی برق" in event["raw_text"]

    def test_raw_cells_dumped(self, tmp_path):
        wb = self._build_sheet()
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = parse_workbook(path)

        assert len(result.raw_cells) == 1
        dump = result.raw_cells[0]
        assert dump["sheet_name"] == "01_05"
        assert dump["report_date"] == date(2026, 3, 25)
        assert dump["cells"]["A1"] == "تاریخ"

    def test_missing_date_records_failure(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "garbage"  # no MM_DD pattern
        ws["A1"] = "no date here"
        path = tmp_path / "bad.xlsx"
        wb.save(path)

        result = parse_workbook(path)
        assert result.sheets_parsed == 0
        assert result.sheets_failed == 1
        assert any("could not find date" in e for e in result.errors)

    def test_falls_back_to_sheet_name_for_date(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "01_05"  # MM_DD format
        ws["A1"] = "no date label here"
        path = tmp_path / "fallback.xlsx"
        wb.save(path)

        result = parse_workbook(path)
        assert result.sheets_parsed == 1
        assert result.daily_reports[0]["report_date"] == date(2026, 3, 25)
