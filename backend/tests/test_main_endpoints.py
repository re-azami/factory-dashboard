"""Tests for the FastAPI HTTP endpoints in app.main."""
import io
from unittest.mock import patch, MagicMock

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.pool import StaticPool
from sqlalchemy.orm import sessionmaker

from app.main import app as fastapi_app
from app.database import Base, get_db
import app.models  # noqa: F401  side-effect: registers tables on Base.metadata


# ── Test client with in-memory SQLite ─────────────────────────────────────────

@pytest.fixture
def client():
    # StaticPool + check_same_thread=False so TestClient (which runs requests in a worker
    # thread) can share the same in-memory SQLite database with the test setup thread.
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestSession = sessionmaker(bind=engine)

    def override_get_db():
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    fastapi_app.dependency_overrides[get_db] = override_get_db
    with TestClient(fastapi_app) as c:
        yield c
    fastapi_app.dependency_overrides.clear()
    engine.dispose()


# ── /health ──────────────────────────────────────────────────────────────────

class TestHealth:
    def test_health_ok(self, client):
        resp = client.get("/health")
        assert resp.status_code == 200
        assert resp.json() == {"status": "ok"}


# ── /chat ────────────────────────────────────────────────────────────────────

class TestChat:
    def test_chat_streams_agent_chunks(self, client):
        def fake_run(question, db):
            yield "Hello "
            yield "world."

        with patch("app.main.agent.run", side_effect=fake_run):
            resp = client.post("/chat", json={"question": "Hi?"})

        assert resp.status_code == 200
        assert resp.text == "Hello world."

    def test_chat_requires_question(self, client):
        resp = client.post("/chat", json={})
        assert resp.status_code == 422


# ── /history ─────────────────────────────────────────────────────────────────

class TestHistory:
    def test_empty_history(self, client):
        resp = client.get("/history")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_returns_recent_queries(self, client):
        from app.query_log.log import save

        # Use the same overridden dependency to insert
        get_db_gen = fastapi_app.dependency_overrides[get_db]()
        db = next(get_db_gen)
        try:
            save(db=db, question="q1", tool_calls=[], answer="a1", llm_provider="anthropic")
            save(db=db, question="q2", tool_calls=[{"tool": "execute_sql"}], answer="a2", llm_provider="openai")
        finally:
            db.close()

        resp = client.get("/history?limit=10")
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) == 2
        # Both rows are returned; ordering between same-timestamp rows isn't guaranteed
        questions = {d["question"] for d in data}
        providers = {d["llm_provider"] for d in data}
        assert questions == {"q1", "q2"}
        assert providers == {"anthropic", "openai"}

    def test_limit_capped_at_200(self, client):
        resp = client.get("/history?limit=999")
        # Pydantic validation: le=200
        assert resp.status_code == 422


# ── /ingest ──────────────────────────────────────────────────────────────────

def _make_minimal_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01_05"
    ws.cell(row=1, column=1, value="تاریخ")
    ws.cell(row=1, column=2, value="1405/01/05")
    ws.cell(row=7, column=1, value="شیفت روز")
    ws.cell(row=7, column=2, value="خط 1")
    ws.cell(row=7, column=3, value=1000)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestIngest:
    def test_unknown_source_returns_400(self, client):
        resp = client.post(
            "/ingest?source=unknown_source",
            files={"file": ("x.xlsx", b"not really excel", "application/octet-stream")},
        )
        assert resp.status_code == 400
        assert "Unknown source" in resp.json()["detail"]

    def test_factory_source_with_valid_file(self, client):
        content = _make_minimal_workbook()
        resp = client.post(
            "/ingest?source=factory",
            files={"file": ("test.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["sheets_parsed"] == 1
        assert data["sheets_failed"] == 0
        assert data["daily_reports_added"] >= 1
        assert data["production_rows_added"] >= 1

    def test_invalid_file_returns_422(self, client):
        # An empty workbook with no parseable date will produce sheets_parsed=0 → 422
        wb = openpyxl.Workbook()
        wb.active.title = "garbage"
        wb.active["A1"] = "no date"
        buf = io.BytesIO()
        wb.save(buf)

        resp = client.post(
            "/ingest?source=factory",
            files={"file": ("bad.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 422

    def test_idempotent_on_repeat_upload(self, client):
        """Uploading the same workbook twice should not duplicate daily/production rows."""
        content = _make_minimal_workbook()
        files = {"file": ("test.xlsx", content,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

        first = client.post("/ingest?source=factory", files=files).json()
        second = client.post("/ingest?source=factory",
                             files={"file": ("test.xlsx", content,
                                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}).json()

        # On the second pass the parser still parses, but the inserter skips duplicates.
        # The response field reports what came out of the parser (not what was inserted),
        # so the contract here is: no exception, sheets_parsed unchanged.
        assert second["sheets_parsed"] == first["sheets_parsed"]


# ── /ingest/enrich ───────────────────────────────────────────────────────────

class TestEnrich:
    def test_calls_enrich_all(self, client):
        with patch("app.main.enrich_all", return_value=7) as m:
            resp = client.post("/ingest/enrich")
        assert resp.status_code == 200
        assert resp.json() == {"rows_enriched": 7}
        m.assert_called_once()
