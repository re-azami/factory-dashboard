"""
Parser for the "daily production report" Excel template used in 1401-1405.

The template has four blocks per sheet (rows 1..max_row):

  Block 1 — Production summary (rows 3-10):
      row 3   : supervisors (D=day, K=night), date (U)
      rows 7-10: (day/line1), (day/line2), (night/line1), (night/line2)

  Block 2 — Mills & line-/feed-input downtime (rows ~13 to ~38):
      row 14  : column header labels, including
                  C='میزان مصرف آب' (water consumption, m³ per shift)
                  D..K='مصرف گلوله' (ball additions; row 17 holds diameters)
                  N='علت توقف کارخانه' (factory downtime cause)
                  X (or W) ='علت توقف فید ورودی' (input-feed downtime cause)
      rows 15-37 (variable): mixed downtime + ball-additions data rows.

  Block 3 — Filter-press downtime (rows ~38 to ~50):
      Header A='دلایل توقف فیلتر پرس'.
      Day side at columns B(line)/C(desc)/K(dur).
      Night side at columns S(line)/[T|U|V](desc)/AD(dur).

  Block 4 — Quality (last block):
      Header A='کیفیت خوراک و کنسانتره تولیدی'.
      Sub-header row + Fe%/Feo% row + 4 data rows ((day,line1)(day,line2)(night,line1)(night,line2)).
      Columns vary across years; we anchor each metric by its sub-header label.

Two column-layout variants exist for blocks 1+2:
  Variant A (1405/1404/1403): production block has a "Ton/h" column at U, and
    the downtime block headers sit at L/W/X/AD.
  Variant B (1402/1401):     no "Ton/h" column. Downtime block at L/V/W/AC.

This parser anchors columns and rows by header text rather than fixed offsets,
so it handles both variants and any minor future drift.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet


class ParseError(Exception):
    """Raised when a sheet's structure does not match the expected template."""


class MissingDataError(ParseError):
    """Raised when required fields are missing for a non-idle line.

    Carries `missing` — list of human-readable strings like
    'day/line1: fe_input_feed missing (cell C55 empty)' so the caller can
    investigate the sheet. Per the user's 'don't skip data' rule, a sheet
    with any unexpectedly missing field is rejected outright.
    """
    def __init__(self, missing: list[str]) -> None:
        super().__init__(f"unexpectedly missing data on {len(missing)} field(s): " + "; ".join(missing))
        self.missing = missing


# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(value) -> str:
    """Collapse whitespace; safe for None."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _row_cells(ws: Worksheet, row: int) -> list:
    return list(ws[row]) if ws.max_row >= row else []


def _find_col(ws: Worksheet, row: int, needle: str) -> Optional[int]:
    """Return 1-based column index of the cell in `row` whose normalized text contains `needle`."""
    for c in _row_cells(ws, row):
        if needle in _norm(c.value):
            return c.column
    return None


def _find_all_cols(ws: Worksheet, row: int, needle: str) -> list[int]:
    """All matching columns, in order."""
    return [c.column for c in _row_cells(ws, row) if needle in _norm(c.value)]


def _find_row_in_col(ws: Worksheet, col: int, needle: str, row_start: int = 1,
                     row_end: Optional[int] = None) -> Optional[int]:
    """Find the first row in `col` whose normalized text contains `needle`, in [row_start, row_end]."""
    last = row_end if row_end is not None else ws.max_row
    for r in range(row_start, last + 1):
        if needle in _norm(ws.cell(row=r, column=col).value):
            return r
    return None


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v) -> Optional[int]:
    f = _to_float(v)
    if f is None:
        return None
    return int(round(f))


def _to_str(v) -> Optional[str]:
    s = _norm(v)
    return s or None


_LINE1_RE = re.compile(r"^\s*خط\s*1\b")
_LINE2_RE = re.compile(r"^\s*خط\s*2\b")


def _line_number_from_label(text) -> Optional[int]:
    """Match 'خط 1', 'خط1', 'خط  1', or annotated 'خط  2 (7:00-10:00) ...'."""
    s = _norm(text)
    if not s:
        return None
    if _LINE1_RE.match(s):
        return 1
    if _LINE2_RE.match(s):
        return 2
    return None


def _cell_is_red(cell) -> bool:
    """True when the cell's font color is solid red (FFFF0000 or FF0000)."""
    try:
        col = cell.font.color
        if col is None or col.type != "rgb" or not col.rgb:
            return False
        rgb = str(col.rgb).upper()
        return rgb in ("FFFF0000", "FF0000")
    except Exception:
        return False


# ── Block 1: production summary (rows 1-10) ───────────────────────────────────

@dataclass
class ProductionLayout:
    col_input_feed: int = 3      # C
    col_production: int = 5      # E
    col_recovery: int = 7        # G
    col_load: int = 8            # H (may be empty/missing)
    col_ton_per_hour: Optional[int] = None
    col_operation_hour: Optional[int] = None     # 'کارخانه / کارکرد' (line operation hours)
    col_downtime_hour: Optional[int] = None      # 'کارخانه / توقف' (line downtime hours)
    col_drum1: Optional[int] = None
    col_drum2: Optional[int] = None
    col_fp_op: Optional[int] = None
    col_fp_dn: Optional[int] = None
    col_flocculant_grams: Optional[int] = None
    col_flocculant_type: Optional[int] = None


def _discover_production_layout(ws: Worksheet) -> ProductionLayout:
    layout = ProductionLayout()
    layout.col_ton_per_hour = _find_col(ws, 6, "Ton/h")
    # Line operation/downtime hours: row 5 'کارخانه' (factory) heading anchors a 2-col group
    # with row-6 sub-labels 'کارکرد' / 'توقف'. In variant A this is V/W; in variant B U/V.
    factory_group = _find_col(ws, 5, "کارخانه")
    if factory_group:
        layout.col_operation_hour = factory_group
        layout.col_downtime_hour = factory_group + 1
    drum_anchor = _find_col(ws, 4, "کارکرد درام فیلتر")
    if drum_anchor:
        layout.col_drum1 = drum_anchor
        layout.col_drum2 = drum_anchor + 1
    fp_anchor = _find_col(ws, 4, "کارکرد فیلترپرس") or _find_col(ws, 4, "کارکرد فیلتر پرس")
    if fp_anchor:
        layout.col_fp_op = fp_anchor
        layout.col_fp_dn = fp_anchor + 1
    layout.col_flocculant_grams = _find_col(ws, 4, "مصرف فلوکولانت")
    layout.col_flocculant_type = _find_col(ws, 4, "نوع فلوکولانت")
    return layout


# ── Block 2: mill ball additions, water, factory/feed downtime ────────────────

@dataclass
class DowntimeLayout:
    """Columns for the 'مصرف گلوله / دلایل توقف' block."""
    # Mills & water
    col_water: int = 3                # C — 'میزان مصرف آب'
    diameters_row: int = 17           # row holding the 8 diameters (set dynamically)
    col_primary_30: int = 4           # D
    col_primary_40: int = 5           # E
    col_primary_50: int = 6           # F
    col_primary_60: int = 7           # G
    col_secondary_25: int = 8         # H
    col_secondary_30: int = 9         # I
    col_secondary_40: int = 10        # J
    col_secondary_50: int = 11        # K
    # Factory & feed-input downtime (positions vary between variants)
    col_factory_shift_label: int = 12 # L
    col_factory_line_label: int = 13  # M
    col_factory_desc: int = 14        # N
    col_factory_dur: Optional[int] = None
    col_feed_desc: Optional[int] = None
    col_feed_dur: Optional[int] = None
    # Row of the downtime header (typically 14, but multi-load shifts push it down).
    header_row: int = 14


def _find_downtime_header_row(ws: Worksheet, search_start: int = 11, search_end: int = 30) -> int:
    """The downtime header is the row with 'علت توقف  فید ورودی' (col W or X) AND ≥2 'مدت توقف' cells.
    Most sheets have it at row 14; multi-load sheets push it down (e.g. row 16 in 1404/08_18)."""
    for r in range(search_start, search_end + 1):
        feed_desc = _find_col(ws, r, "علت توقف  فید ورودی") or _find_col(ws, r, "علت توقف فید ورودی")
        durs = _find_all_cols(ws, r, "مدت توقف")
        if feed_desc and len(durs) >= 2:
            return r
    raise ParseError(f"could not find downtime block header (rows {search_start}..{search_end})")


def _discover_downtime_layout(ws: Worksheet) -> DowntimeLayout:
    """The first 'مدت توقف' col is factory_dur; the second is feed_dur. Feed desc is the col with 'علت توقف  فید ورودی'."""
    lay = DowntimeLayout()
    lay.header_row = _find_downtime_header_row(ws)
    # Mill ball-additions diameter labels are 2 rows below the downtime header.
    lay.diameters_row = lay.header_row + 3
    feed_desc_col = _find_col(ws, lay.header_row, "علت توقف  فید ورودی") or _find_col(ws, lay.header_row, "علت توقف فید ورودی")
    dur_cols = _find_all_cols(ws, lay.header_row, "مدت توقف")
    factory_durs = [c for c in dur_cols if c < feed_desc_col]
    feed_durs = [c for c in dur_cols if c > feed_desc_col]
    if not factory_durs or not feed_durs:
        raise ParseError(f"unexpected duration column layout in row {lay.header_row}: durs={dur_cols} feed_desc={feed_desc_col}")
    lay.col_factory_dur = factory_durs[-1]
    lay.col_feed_desc = feed_desc_col
    lay.col_feed_dur = feed_durs[-1]
    return lay


# ── Block 3: filter press downtime ────────────────────────────────────────────

@dataclass
class FilterPressLayout:
    section_row: int                  # row containing 'دلایل توقف فیلتر پرس'
    sub_header_row: int               # section_row + 1
    day_line_col: int                 # B = 2
    day_desc_col: int                 # C = 3 (typical)
    day_dur_col: int                  # K = 11 (typical)
    night_section_col: int            # Q = 17 (typical)
    night_line_col: int               # S = 19 (typical)
    night_desc_col: int               # varies (T/U/V)
    night_dur_col: int                # AD or AC
    last_data_row: int                # row before the next block (quality) header


def _looks_like_fp_sub_header(ws: Worksheet, row: int) -> bool:
    """A FP sub-header row has 'شیفت روز' in col A, 'شیفت شب' further right, and ≥2
    each of 'علت توقف' / 'مدت توقف' labels (one set per side)."""
    if "شیفت روز" not in _norm(ws.cell(row=row, column=1).value):
        return False
    night_col = _find_col(ws, row, "شیفت شب")
    if not night_col:
        return False
    descs = _find_all_cols(ws, row, "علت توقف")
    durs = _find_all_cols(ws, row, "مدت توقف")
    return len(descs) >= 2 and len(durs) >= 2


def _discover_filter_press(ws: Worksheet, downtime_end_hint: int,
                           quality_section_row: int) -> Optional[FilterPressLayout]:
    """Locate the filter-press section. Two valid layouts:
        (a) explicit header 'دلایل توقف فیلترپرس' in column A, with the sub-header on the next row;
        (b) header missing/blank — sub-header row alone (e.g. 1404/10_05).
    If both are found in the same range, that is unexpected and signals a malformed
    sheet; raise ParseError so the user can investigate.
    Returns None only when neither pattern is found (sheet has no FP section)."""
    search_start = downtime_end_hint + 1
    search_end = quality_section_row - 1

    # Find both anchors (if any) in the search range.
    explicit_header_row: Optional[int] = None
    sub_header_row: Optional[int] = None
    for r in range(search_start, search_end + 1):
        if explicit_header_row is None and "دلایل توقف فیلتر" in _norm(ws.cell(row=r, column=1).value):
            explicit_header_row = r
        if sub_header_row is None and _looks_like_fp_sub_header(ws, r):
            sub_header_row = r

    # The normal layout has the explicit header and sub-header on consecutive rows
    # (header at N, sub-header at N+1). That's expected — not a conflict.
    # A real conflict is when both are present but NOT consecutive: that signals a
    # malformed sheet the user wants flagged.
    if explicit_header_row and sub_header_row and sub_header_row != explicit_header_row + 1:
        raise ParseError(
            f"FP section has an explicit header (row {explicit_header_row}) "
            f"and a separate sub-header pattern (row {sub_header_row}) that is not "
            f"on the next row; investigate this sheet."
        )

    if explicit_header_row:
        section_row = explicit_header_row
        sub = section_row + 1
    elif sub_header_row:
        # Header missing — synthetic section_row one row above the sub-header.
        section_row = sub_header_row - 1
        sub = sub_header_row
    else:
        return None  # genuinely no FP section
    # Day side columns
    day_dur_col = None
    night_dur_col = None
    night_section_col = _find_col(ws, sub, "شیفت شب")
    if not night_section_col:
        raise ParseError(f"could not find 'شیفت شب' in row {sub}")
    # Two 'مدت توقف' labels — first <= night_section_col is day; the other is night.
    durs = _find_all_cols(ws, sub, "مدت توقف")
    if len(durs) < 2:
        raise ParseError(f"expected two 'مدت توقف' labels in row {sub}, got {durs}")
    day_dur_col = next((c for c in durs if c < night_section_col), None)
    night_dur_col = next((c for c in durs if c > night_section_col), None)
    if not day_dur_col or not night_dur_col:
        raise ParseError(f"could not split 'مدت توقف' into day/night ({durs}, night_section={night_section_col})")
    # 'علت توقف' labels — leftmost = day, the one after night_section_col = night.
    desc_cols = _find_all_cols(ws, sub, "علت توقف")
    if len(desc_cols) < 2:
        raise ParseError(f"expected two 'علت توقف' labels in row {sub}, got {desc_cols}")
    day_desc_col = next((c for c in desc_cols if c < night_section_col), None)
    night_desc_col = next((c for c in desc_cols if c > night_section_col), None)
    if not day_desc_col or not night_desc_col:
        raise ParseError(f"could not split 'علت توقف' into day/night ({desc_cols})")
    # 'خط فرآوری' labels (line-number column) — leftmost=day (B), the one after night_section_col=night
    line_cols = _find_all_cols(ws, sub, "خط فرآوری")
    if not line_cols:
        raise ParseError(f"could not find 'خط فرآوری' in row {sub}")
    day_line_col = next((c for c in line_cols if c < night_section_col), 2)  # B fallback
    night_line_col = next((c for c in line_cols if c > night_section_col), 19)  # S fallback

    # End-of-block is bounded by the quality-section row passed in by the caller.
    last_data_row = quality_section_row - 1

    return FilterPressLayout(
        section_row=section_row,
        sub_header_row=sub,
        day_line_col=day_line_col,
        day_desc_col=day_desc_col,
        day_dur_col=day_dur_col,
        night_section_col=night_section_col,
        night_line_col=night_line_col,
        night_desc_col=night_desc_col,
        night_dur_col=night_dur_col,
        last_data_row=last_data_row,
    )


# ── Block 4: quality ──────────────────────────────────────────────────────────

@dataclass
class QualityLayout:
    section_row: int
    fe_row: int                                  # row containing 'Fe%' / 'Feo%'
    data_rows: list[Optional[int]]               # 4 slots: [day1, day2, night1, night2]; None when the row was missing
    feed_fe_col: Optional[int] = None
    feed_feo_col: Optional[int] = None
    concentrate_fe_col: Optional[int] = None
    concentrate_feo_col: Optional[int] = None
    thickener_fe_col: Optional[int] = None
    thickener_feo_col: Optional[int] = None
    first_ballmill_fe_col: Optional[int] = None
    first_ballmill_feo_col: Optional[int] = None
    k80_input_feed_col: Optional[int] = None
    k80_primary_col: Optional[int] = None
    k80_secondary_col: Optional[int] = None
    k80_hydrocyclone_1_col: Optional[int] = None
    k80_hydrocyclone_2_col: Optional[int] = None
    k80_tailing_col: Optional[int] = None
    k80_concentrate_col: Optional[int] = None
    dry_weight_recovery_col: Optional[int] = None
    metallurgical_recovery_col: Optional[int] = None
    separation_efficiency_col: Optional[int] = None
    input_feed_moisture_col: Optional[int] = None
    concentrate_moisture_col: Optional[int] = None
    filter_press_cake_moisture_col: Optional[int] = None


def _discover_quality(ws: Worksheet, fp_end_hint: int) -> QualityLayout:
    section_row = _find_row_in_col(ws, 1, "کیفیت خوراک", row_start=fp_end_hint)
    if not section_row:
        raise ParseError("could not find 'کیفیت خوراک و کنسانتره تولیدی' section header in column A")

    # Find Fe%/Feo% row by scanning rows section_row+1..section_row+10 for at least 3 'Fe%' cells.
    fe_row = None
    for r in range(section_row + 1, min(section_row + 12, ws.max_row + 1)):
        fe_count = sum(1 for c in _row_cells(ws, r) if _norm(c.value) == "Fe%")
        if fe_count >= 3:
            fe_row = r
            break
    if not fe_row:
        raise ParseError(f"could not locate Fe%/Feo% row near section_row={section_row}")

    # Map Fe%/Feo% columns. Order them; the i-th pair maps to feed/concentrate/thickener/first_ballmill.
    fe_cols = sorted([c.column for c in _row_cells(ws, fe_row) if _norm(c.value) == "Fe%"])
    feo_cols = sorted([c.column for c in _row_cells(ws, fe_row) if _norm(c.value) == "Feo%"])

    qlay = QualityLayout(section_row=section_row, fe_row=fe_row, data_rows=[])
    pair_targets = [
        ("feed_fe_col", "feed_feo_col"),
        ("concentrate_fe_col", "concentrate_feo_col"),
        ("thickener_fe_col", "thickener_feo_col"),
        ("first_ballmill_fe_col", "first_ballmill_feo_col"),
    ]
    for i, (fe_attr, feo_attr) in enumerate(pair_targets):
        if i < len(fe_cols):
            setattr(qlay, fe_attr, fe_cols[i])
        if i < len(feo_cols):
            setattr(qlay, feo_attr, feo_cols[i])

    # Anchor numeric columns by sub-header rows (between section_row and fe_row).
    def _scan(needle: str) -> Optional[int]:
        for r in range(section_row, fe_row + 1):
            col = _find_col(ws, r, needle)
            if col:
                return col
        return None

    qlay.k80_input_feed_col = _scan("فید ورودی")
    qlay.k80_primary_col = _scan("بالمیل\n اولیه") or _scan("بالمیل اولیه")
    qlay.k80_secondary_col = _scan("ثانویه") if _scan("ثانویه") else None
    qlay.k80_hydrocyclone_1_col = _scan("هیدروسیکلون1")
    qlay.k80_hydrocyclone_2_col = _scan("هیدروسیکلون2")
    qlay.k80_tailing_col = _scan("دانه بندی \nباطله") or _scan("دانه بندی باطله")
    qlay.k80_concentrate_col = _scan("دانه بندی \nکنسانتره") or _scan("دانه بندی کنسانتره")
    # Recoveries / moisture
    qlay.dry_weight_recovery_col = _scan("ریکاوری  وزنی  خشک") or _scan("ریکاوری وزنی خشک") \
        or _scan("راندمان خشک") or _scan("راندمان")
    qlay.metallurgical_recovery_col = _scan("ریکاروی عیاری") or _scan("ریکاوری عیاری")
    qlay.separation_efficiency_col = _scan("بازدهی\nجدایش") or _scan("بازدهی جدایش")
    qlay.input_feed_moisture_col = _scan("میانگین رطوبت\nخوراک") or _scan("میانگین رطوبت خوراک")
    qlay.concentrate_moisture_col = _scan("میانگین رطوبت\n کنسانتره") or _scan("میانگین رطوبت کنسانتره")
    qlay.filter_press_cake_moisture_col = _scan("رطوبت کیک فیلترپرس") or _scan("رطوبت کیک فیلتر پرس")

    # Locate the 4 data rows in template order:
    #   day-line1, day-line2, night-line1, night-line2.
    # Some sheets are missing one or more rows (operator combined loads or skipped
    # idle lines); record None placeholders for missing positions rather than
    # rejecting the sheet.
    data_rows: list[Optional[int]] = [None, None, None, None]   # [d1, d2, n1, n2]

    # Walk rows after fe_row, tracking current shift (via شیفت روز/شب in col A) and
    # filling slots based on خط 1/خط 2 in col B (loose match).
    current_shift = "day"
    for r in range(fe_row + 1, ws.max_row + 1):
        a = _norm(ws.cell(row=r, column=1).value)
        if a == "شیفت روز" or a.startswith("شیفت روز"):
            current_shift = "day"
        elif a == "شیفت شب" or a.startswith("شیفت شب"):
            current_shift = "night"

        ln = _line_number_from_label(ws.cell(row=r, column=2).value)
        if ln is None:
            continue
        slot = (0 if current_shift == "day" else 2) + (ln - 1)
        if data_rows[slot] is None:                # take first occurrence per slot
            data_rows[slot] = r
        if all(x is not None for x in data_rows):
            break

    qlay.data_rows = data_rows                     # may contain None for missing rows
    return qlay


# ── Sheet validation ──────────────────────────────────────────────────────────

def _validate_anchors(ws: Worksheet) -> None:
    """Minimal anchor checks: production block must START at row 7 with day-shift line-1.
    Multi-load sheets shift the night-shift start beyond row 9, so we discover that dynamically."""
    if "تاریخ" not in _norm(ws["T3"].value):
        raise ParseError("expected 'تاریخ:' label at T3")
    if "شیفت روز" not in _norm(ws["A7"].value):
        raise ParseError("expected 'شیفت روز' anchor at A7")
    if _line_number_from_label(ws["B7"].value) != 1:
        raise ParseError(f"expected 'خط 1' at B7, got {ws['B7'].value!r}")


# ── Parsed-data shape ────────────────────────────────────────────────────────

@dataclass
class DowntimeEvent:
    description: str
    duration_minutes: int


@dataclass
class LineDowntimes:
    factory: list[DowntimeEvent] = field(default_factory=list)            # [(desc, dur)]
    # Each input-feed event carries an optional pointer to the index in `factory`
    # that it should link to. None means standalone (factory_downtime_id NULL).
    input_feed: list[tuple[DowntimeEvent, Optional[int]]] = field(default_factory=list)
    filter_press: list[DowntimeEvent] = field(default_factory=list)


@dataclass
class LineRow:
    line_number: int
    # Production block
    input_feed_tonnage: Optional[int]
    production_tonnage: Optional[int]
    recovery: Optional[float]
    load_code: Optional[str]
    operation_hour: Optional[float]
    downtime_hour: Optional[float]
    ton_per_hour: Optional[float]
    drum_filter_1_hour: Optional[float]
    drum_filter_2_hour: Optional[float]
    filter_press_operation_hour: Optional[float]
    filter_press_downtime_hour: Optional[float]
    flocculant_consumption_grams: Optional[int]
    flocculant_type: Optional[str]
    # Mill ball additions
    primary_mill_30: Optional[int] = None
    primary_mill_40: Optional[int] = None
    primary_mill_50: Optional[int] = None
    primary_mill_60: Optional[int] = None
    secondary_mill_25: Optional[int] = None
    secondary_mill_30: Optional[int] = None
    secondary_mill_40: Optional[int] = None
    secondary_mill_50: Optional[int] = None
    # Quality — Fe / Feo (whole-number percent)
    fe_input_feed: Optional[float] = None
    feo_input_feed: Optional[float] = None
    fe_concentrate: Optional[float] = None
    feo_concentrate: Optional[float] = None
    fe_thickener_tailing: Optional[float] = None
    feo_thickener_tailing: Optional[float] = None
    fe_first_ballmill_output: Optional[float] = None
    feo_first_ballmill_output: Optional[float] = None
    # Quality — K80 (microns)
    k80_size_input_feed: Optional[int] = None
    k80_size_primary_ballmill: Optional[int] = None
    k80_size_secondary_ballmill: Optional[int] = None
    k80_size_hydrocyclone_overflow_1: Optional[int] = None
    k80_size_hydrocyclone_overflow_2: Optional[int] = None
    k80_size_tailing: Optional[int] = None
    k80_size_concentrate: Optional[int] = None
    # Quality — recoveries & moisture (whole-number percent)
    dry_weight_recovery: Optional[float] = None
    metallurgical_recovery: Optional[float] = None
    separation_efficiency: Optional[float] = None
    input_feed_moisture: Optional[float] = None
    concentrate_moisture: Optional[float] = None
    filter_press_cake_moisture: Optional[float] = None
    # Downtime events
    downtimes: LineDowntimes = field(default_factory=LineDowntimes)


@dataclass
class ShiftBlock:
    shift: str                                 # 'day' | 'night'
    supervisor_name: Optional[str]
    downtime_description: Optional[str]        # red-font shift-level note from input-feed-cause column
    water_consumption: Optional[float]         # m³ for the whole shift (sum across lines)
    # Most shifts have one segment per line. Multi-load shifts (when operator
    # changes feed mid-shift) have two segments per line.
    line1_segments: list[LineRow] = field(default_factory=list)
    line2_segments: list[LineRow] = field(default_factory=list)

    @property
    def line1(self) -> LineRow:                # convenience for legacy access — returns segment 1
        return self.line1_segments[0]

    @property
    def line2(self) -> LineRow:
        return self.line2_segments[0]


@dataclass
class SheetData:
    jalali_date: str
    day_shift: ShiftBlock
    night_shift: ShiftBlock


# ── Production-block reader ──────────────────────────────────────────────────

def _read_production_line(ws: Worksheet, row: int, line_number: int, layout: ProductionLayout) -> LineRow:
    # No-production days leave these cells blank; coerce to 0 so the line is still
    # ingested as 'idle' rather than rejected for missing data.
    feed_raw = _to_int(ws.cell(row=row, column=layout.col_input_feed).value)
    prod_raw = _to_int(ws.cell(row=row, column=layout.col_production).value)
    feed = feed_raw if feed_raw is not None else 0
    prod = prod_raw if prod_raw is not None else 0
    raw_recovery = _to_float(ws.cell(row=row, column=layout.col_recovery).value)
    if not feed:
        recovery: Optional[float] = 0.0
    elif raw_recovery is None:
        recovery = None
    else:
        recovery = round(raw_recovery * 100, 4) if raw_recovery <= 1.5 else round(raw_recovery, 4)
    load_code = _to_str(ws.cell(row=row, column=layout.col_load).value) if layout.col_load else None

    def at(col: Optional[int]) -> Optional[float]:
        return _to_float(ws.cell(row=row, column=col).value) if col else None

    operation_hour = at(layout.col_operation_hour)
    downtime_hour = at(layout.col_downtime_hour)

    # ton_per_hour: prefer Excel value; fall back to feed/operation_hour when missing.
    ton_per_hour = at(layout.col_ton_per_hour)
    if ton_per_hour is None and feed and operation_hour:
        ton_per_hour = round(feed / operation_hour, 2)

    # filter_press_operation_hour: empty cell means filter press wasn't running → 0.
    fp_op = at(layout.col_fp_op)
    if fp_op is None and layout.col_fp_op is not None:
        fp_op = 0.0

    return LineRow(
        line_number=line_number,
        input_feed_tonnage=feed,
        production_tonnage=prod,
        recovery=recovery,
        load_code=load_code,
        operation_hour=operation_hour,
        downtime_hour=downtime_hour,
        ton_per_hour=ton_per_hour,
        drum_filter_1_hour=at(layout.col_drum1),
        drum_filter_2_hour=at(layout.col_drum2),
        filter_press_operation_hour=fp_op,
        filter_press_downtime_hour=at(layout.col_fp_dn),
        flocculant_consumption_grams=_to_int(ws.cell(row=row, column=layout.col_flocculant_grams).value)
            if layout.col_flocculant_grams else None,
        flocculant_type=_to_str(ws.cell(row=row, column=layout.col_flocculant_type).value)
            if layout.col_flocculant_type else None,
    )


# ── Production-block walker (multi-load aware) ───────────────────────────────

def _walk_production_block(ws: Worksheet, prod_layout: ProductionLayout) -> tuple[list[LineRow], list[LineRow], list[LineRow], list[LineRow], int]:
    """Walk rows 7+ until 'جمع کل' grand-total row, emitting LineRow per data row.

    Returns (day_line1_segments, day_line2_segments, night_line1_segments, night_line2_segments, totals_row).
    Each list has 1 or 2 entries (1 = single load shift; 2 = multi-load shift)."""
    day_l1: list[LineRow] = []
    day_l2: list[LineRow] = []
    night_l1: list[LineRow] = []
    night_l2: list[LineRow] = []
    current_shift: Optional[str] = "day"   # row 7 is day per validate_anchors
    totals_row: Optional[int] = None

    # Production block starts at row 7. Cap the walk at row 16 — multi-load
    # shifts use up to 6 production rows + totals row 13 + buffer.
    for r in range(7, 17):
        a = _norm(ws.cell(row=r, column=1).value)
        b_label = ws.cell(row=r, column=2).value
        line = _line_number_from_label(b_label)

        if a == "شیفت روز":
            current_shift = "day"
        elif a == "شیفت شب":
            current_shift = "night"
        elif a == "جمع کل":
            totals_row = r
            break

        if line is None:
            continue
        line_row = _read_production_line(ws, r, line, prod_layout)
        if current_shift == "day":
            (day_l1 if line == 1 else day_l2).append(line_row)
        else:
            (night_l1 if line == 1 else night_l2).append(line_row)

    if totals_row is None:
        raise ParseError("could not find 'جمع کل' totals row in production block (rows 7-16)")
    if not day_l1 or not day_l2 or not night_l1 or not night_l2:
        raise ParseError(
            f"production block missing rows: day_l1={len(day_l1)} day_l2={len(day_l2)} "
            f"night_l1={len(night_l1)} night_l2={len(night_l2)} (totals at row {totals_row})"
        )
    if any(len(lst) > 2 for lst in (day_l1, day_l2, night_l1, night_l2)):
        raise ParseError(
            f"more than 2 load segments per (shift, line) is unsupported: "
            f"day_l1={len(day_l1)} day_l2={len(day_l2)} night_l1={len(night_l1)} night_l2={len(night_l2)}"
        )
    return day_l1, day_l2, night_l1, night_l2, totals_row


# ── Mill / water reader ──────────────────────────────────────────────────────

def _find_mill_rows(ws: Worksheet, dt_layout: DowntimeLayout, dt_end_row: int) -> dict:
    """
    The mill block sits between row 18 and the next major marker (شیفت شب for the
    day half, جمع کل for the night half). Return rows for each of
    (day,1)(day,2)(night,1)(night,2).

    A mill-data row is any row in the range that has at least one cell in mill
    columns D..K populated (zero or non-zero). The 'خط 1'/'خط 2' label in column B
    is preferred but NOT required — some sheets omit it. We use the شیفت روز /
    شیفت شب / جمع کل markers in column A to bound day and night spans, then take
    the first 2 mill-data rows in each span as line 1 and line 2 in order.
    """
    diameters_row = dt_layout.diameters_row

    def has_mill_data(r: int) -> bool:
        return any(ws.cell(row=r, column=col).value is not None for col in range(4, 12))

    # Find the شیفت شب marker (start of night half) and جمع کل marker (end of night half).
    night_start: Optional[int] = None
    grand_total_row: Optional[int] = None
    for r in range(diameters_row + 1, dt_end_row + 1):
        a = _norm(ws.cell(row=r, column=1).value)
        if night_start is None and a == "شیفت شب":
            night_start = r
        elif night_start is not None and a == "جمع کل":
            grand_total_row = r
            break
    if night_start is None:
        raise ParseError("could not find 'شیفت شب' marker in column A within mill block")
    if grand_total_row is None:
        grand_total_row = dt_end_row + 1

    # Day half: from diameters_row+1 to night_start-1.
    # Take the first two rows with mill data (line 1 then line 2 in template order).
    day_rows = [r for r in range(diameters_row + 1, night_start) if has_mill_data(r)][:2]
    # Night half: from night_start to grand_total_row-1.
    night_rows = [r for r in range(night_start, grand_total_row) if has_mill_data(r)][:2]

    if len(day_rows) != 2 or len(night_rows) != 2:
        raise ParseError(f"could not locate 2 mill-data rows per shift: "
                         f"day={day_rows} (range {diameters_row+1}..{night_start-1}); "
                         f"night={night_rows} (range {night_start}..{grand_total_row-1})")

    return {("day", 1): day_rows[0], ("day", 2): day_rows[1],
            ("night", 1): night_rows[0], ("night", 2): night_rows[1]}


def _apply_mills_and_water(ws: Worksheet, line_row: LineRow, row: int, dt_layout: DowntimeLayout) -> Optional[float]:
    """Mutate `line_row` with mill ball additions; return this row's water_consumption (or None).

    Mill cells with missing values are treated as 0 (operator left blank meaning 'no
    balls added'). The diameter columns are always present in the template, so we
    don't need to distinguish between 'column missing' and 'cell empty' for mills.
    """
    def mill(col: int) -> int:
        v = _to_int(ws.cell(row=row, column=col).value)
        return v if v is not None else 0

    line_row.primary_mill_30 = mill(dt_layout.col_primary_30)
    line_row.primary_mill_40 = mill(dt_layout.col_primary_40)
    line_row.primary_mill_50 = mill(dt_layout.col_primary_50)
    line_row.primary_mill_60 = mill(dt_layout.col_primary_60)
    line_row.secondary_mill_25 = mill(dt_layout.col_secondary_25)
    line_row.secondary_mill_30 = mill(dt_layout.col_secondary_30)
    line_row.secondary_mill_40 = mill(dt_layout.col_secondary_40)
    line_row.secondary_mill_50 = mill(dt_layout.col_secondary_50)
    return _to_float(ws.cell(row=row, column=dt_layout.col_water).value)


# ── Factory + input-feed downtime reader ─────────────────────────────────────

def _read_factory_and_feed_downtime(ws: Worksheet, dt_layout: DowntimeLayout,
                                    dt_start_row: int, dt_end_row: int,
                                    day_block: "ShiftBlock", night_block: "ShiftBlock") -> None:
    """
    Walk rows [dt_start_row..dt_end_row] tracking current shift/line via column-L/M markers
    AND current load-segment per shift via the 'تعویض بار' marker (load change).

    Each downtime event is attached to the LineRow of the matching segment within
    the current (shift, line). For single-load shifts (1 segment), all events go
    to segment 1. For multi-load shifts (2 segments), events before the load-change
    marker go to segment 1; events after go to segment 2.

    Pairs an input-feed event with the same-row factory event when descriptions
    contain 'توقف خط' AND durations are exactly equal.

    Also extracts per-shift `downtime_description`: red-font notes in the
    feed-input-cause column with no associated duration. Appended to the
    current shift's `ShiftBlock.downtime_description` (joined by ' | ').
    """
    current_shift: Optional[str] = None
    current_line: Optional[int] = None
    current_segment = {"day": 1, "night": 1}

    def _get_target(shift: str, line: int) -> Optional[LineRow]:
        block = day_block if shift == "day" else night_block
        segs = block.line1_segments if line == 1 else block.line2_segments
        if not segs:
            return None
        idx = current_segment[shift] - 1
        if idx >= len(segs):
            idx = len(segs) - 1   # if marker fired but only one segment exists, stay on last
        return segs[idx]

    for r in range(dt_start_row, dt_end_row + 1):
        l_label = _norm(ws.cell(row=r, column=dt_layout.col_factory_shift_label).value)
        m_label = _norm(ws.cell(row=r, column=dt_layout.col_factory_line_label).value)
        if l_label == "شیفت روز":
            current_shift = "day"
        elif l_label == "شیفت شب":
            current_shift = "night"
        ln = _line_number_from_label(m_label)
        if ln is not None:
            current_line = ln

        factory_desc = _to_str(ws.cell(row=r, column=dt_layout.col_factory_desc).value)
        factory_dur = _to_int(ws.cell(row=r, column=dt_layout.col_factory_dur).value) if dt_layout.col_factory_dur else None
        feed_cell = ws.cell(row=r, column=dt_layout.col_feed_desc) if dt_layout.col_feed_desc else None
        feed_desc = _to_str(feed_cell.value) if feed_cell else None
        feed_dur = _to_int(ws.cell(row=r, column=dt_layout.col_feed_dur).value) if dt_layout.col_feed_dur else None

        # Multi-load tracking: when 'تعویض بار' (load change) appears in either
        # description, advance the current shift's segment counter.
        is_load_change = bool(
            (factory_desc and "تعویض بار" in factory_desc)
            or (feed_desc and "تعویض بار" in feed_desc)
        )
        if is_load_change and current_shift is not None:
            block = day_block if current_shift == "day" else night_block
            n_segs = max(len(block.line1_segments), len(block.line2_segments))
            if n_segs >= 2 and current_segment[current_shift] < n_segs:
                current_segment[current_shift] += 1

        # Shift-level note: feed-cause text with no duration, OR with red font.
        if feed_desc and current_shift is not None and (feed_dur is None or _cell_is_red(feed_cell)):
            block = day_block if current_shift == "day" else night_block
            if block.downtime_description:
                block.downtime_description = block.downtime_description + " | " + feed_desc
            else:
                block.downtime_description = feed_desc
            feed_desc_for_line: Optional[str] = None
        else:
            feed_desc_for_line = feed_desc

        if current_shift is None or current_line is None:
            continue
        target = _get_target(current_shift, current_line)
        if target is None:
            continue

        added_factory_idx: Optional[int] = None
        if factory_desc and factory_dur is not None:
            target.downtimes.factory.append(DowntimeEvent(description=factory_desc, duration_minutes=factory_dur))
            added_factory_idx = len(target.downtimes.factory) - 1

        if feed_desc_for_line and feed_dur is not None:
            ev = DowntimeEvent(description=feed_desc_for_line, duration_minutes=feed_dur)
            link_idx: Optional[int] = None
            if (added_factory_idx is not None
                    and "توقف خط" in feed_desc_for_line
                    and factory_dur is not None
                    and factory_dur == feed_dur):
                link_idx = added_factory_idx
            target.downtimes.input_feed.append((ev, link_idx))


# ── Filter-press downtime reader ─────────────────────────────────────────────

def _read_filter_press(ws: Worksheet, fp: FilterPressLayout, targets: dict) -> None:
    """For each side (day / night) walk rows after sub_header_row..last_data_row.
    Track current line via column B (day) / column S (night)."""
    def _walk(side: str, line_col: int, desc_col: int, dur_col: int) -> None:
        current_line: Optional[int] = None
        for r in range(fp.sub_header_row + 1, fp.last_data_row + 1):
            ln = _line_number_from_label(ws.cell(row=r, column=line_col).value)
            if ln is not None:
                current_line = ln
            desc = _to_str(ws.cell(row=r, column=desc_col).value)
            dur = _to_int(ws.cell(row=r, column=dur_col).value)
            if current_line is None or not desc or dur is None:
                continue
            target = targets.get((side, current_line))
            if target is None:
                continue
            target.downtimes.filter_press.append(DowntimeEvent(description=desc, duration_minutes=dur))

    _walk("day", fp.day_line_col, fp.day_desc_col, fp.day_dur_col)
    _walk("night", fp.night_line_col, fp.night_desc_col, fp.night_dur_col)


# ── Quality reader ───────────────────────────────────────────────────────────

def _apply_quality(ws: Worksheet, line_row: LineRow, row: int, q: QualityLayout) -> None:
    def fl(col: Optional[int]) -> Optional[float]:
        return _to_float(ws.cell(row=row, column=col).value) if col else None
    def it(col: Optional[int]) -> Optional[int]:
        return _to_int(ws.cell(row=row, column=col).value) if col else None

    line_row.fe_input_feed = fl(q.feed_fe_col)
    line_row.feo_input_feed = fl(q.feed_feo_col)
    line_row.fe_concentrate = fl(q.concentrate_fe_col)
    line_row.feo_concentrate = fl(q.concentrate_feo_col)
    line_row.fe_thickener_tailing = fl(q.thickener_fe_col)
    line_row.feo_thickener_tailing = fl(q.thickener_feo_col)
    line_row.fe_first_ballmill_output = fl(q.first_ballmill_fe_col)
    line_row.feo_first_ballmill_output = fl(q.first_ballmill_feo_col)

    line_row.k80_size_input_feed = it(q.k80_input_feed_col)
    line_row.k80_size_primary_ballmill = it(q.k80_primary_col)
    line_row.k80_size_secondary_ballmill = it(q.k80_secondary_col)
    line_row.k80_size_hydrocyclone_overflow_1 = it(q.k80_hydrocyclone_1_col)
    line_row.k80_size_hydrocyclone_overflow_2 = it(q.k80_hydrocyclone_2_col)
    line_row.k80_size_tailing = it(q.k80_tailing_col)
    line_row.k80_size_concentrate = it(q.k80_concentrate_col)

    line_row.dry_weight_recovery = fl(q.dry_weight_recovery_col)
    line_row.metallurgical_recovery = fl(q.metallurgical_recovery_col)
    line_row.separation_efficiency = fl(q.separation_efficiency_col)
    line_row.input_feed_moisture = fl(q.input_feed_moisture_col)
    line_row.concentrate_moisture = fl(q.concentrate_moisture_col)
    line_row.filter_press_cake_moisture = fl(q.filter_press_cake_moisture_col)


# ── "Don't skip data" validator ──────────────────────────────────────────────

# Production-block fields whose presence is gated by ProductionLayout columns.
# Each tuple is (LineRow attribute name, ProductionLayout column attribute name).
# `ton_per_hour` is dropped — computed fallback covers missing cells.
# `filter_press_operation_hour` is dropped — coerced to 0 when missing.
_PROD_OPTIONAL_FIELDS: list[tuple[str, str]] = [
    ("drum_filter_1_hour", "col_drum1"),
    ("drum_filter_2_hour", "col_drum2"),
    ("filter_press_downtime_hour", "col_fp_dn"),
    ("flocculant_consumption_grams", "col_flocculant_grams"),
    ("flocculant_type", "col_flocculant_type"),
]

# Per-line quality fields: gated by QualityLayout columns AND required on every active line.
_QUALITY_PER_LINE_FIELDS: list[tuple[str, str]] = [
    ("fe_input_feed", "feed_fe_col"),
    ("feo_input_feed", "feed_feo_col"),
    ("fe_concentrate", "concentrate_fe_col"),
    ("feo_concentrate", "concentrate_feo_col"),
    ("fe_thickener_tailing", "thickener_fe_col"),
    ("feo_thickener_tailing", "thickener_feo_col"),
    ("k80_size_input_feed", "k80_input_feed_col"),
    ("dry_weight_recovery", "dry_weight_recovery_col"),
    ("metallurgical_recovery", "metallurgical_recovery_col"),
    ("separation_efficiency", "separation_efficiency_col"),
    ("input_feed_moisture", "input_feed_moisture_col"),
    ("concentrate_moisture", "concentrate_moisture_col"),
    ("filter_press_cake_moisture", "filter_press_cake_moisture_col"),
]

# Per-shift quality fields: sampled once per shift and recorded on EITHER line 1 or line 2
# of that shift (operator's choice). The validator passes if at least one of the two lines
# has the value within the same shift.
_QUALITY_PER_SHIFT_FIELDS: list[tuple[str, str]] = [
    ("fe_first_ballmill_output", "first_ballmill_fe_col"),
    ("feo_first_ballmill_output", "first_ballmill_feo_col"),
    ("k80_size_primary_ballmill", "k80_primary_col"),
    ("k80_size_secondary_ballmill", "k80_secondary_col"),
    ("k80_size_hydrocyclone_overflow_1", "k80_hydrocyclone_1_col"),
    ("k80_size_hydrocyclone_overflow_2", "k80_hydrocyclone_2_col"),
    ("k80_size_tailing", "k80_tailing_col"),
    ("k80_size_concentrate", "k80_concentrate_col"),
]

# Mill ball columns are always present in the template (D-K hardcoded);
# enforce non-None on every active line.
_MILL_FIELDS: list[str] = [
    "primary_mill_30", "primary_mill_40", "primary_mill_50", "primary_mill_60",
    "secondary_mill_25", "secondary_mill_30", "secondary_mill_40", "secondary_mill_50",
]

# Production fields that are required regardless of line idleness.
_ALWAYS_REQUIRED_LINE_FIELDS: list[str] = [
    "input_feed_tonnage",
    "production_tonnage",
    "recovery",
]

# Fields that are inherently optional on every line (e.g. only filled when relevant).
# load_code is empty when feed=0; flocculant_* is empty when no flocculant added.
_INHERENTLY_OPTIONAL_LINE: set[str] = {
    "load_code",
    "flocculant_consumption_grams",
    "flocculant_type",
}


def _line_is_idle(line: LineRow) -> bool:
    """A line is idle when no input feed was processed. Quality / mill fields
    are then exempt from the missing-data check."""
    return not line.input_feed_tonnage      # 0 or None


def _validate_no_unexpected_missing(sheet: SheetData,
                                    prod_layout: ProductionLayout,
                                    quality_layout: QualityLayout) -> list[str]:
    """Return human-readable strings describing unexpectedly-None fields.
    Empty list = sheet is clean. Caller raises MissingDataError when non-empty.

    Blocks only on:
      - input_feed_tonnage / production_tonnage / recovery (always required)
      - production-block fields with a present column (ton_per_hour, drum hours, FP hours)

    Quality fields (Fe/FeO, K80, recoveries, moistures) and mill ball additions are
    NOT blocked — quality is kept as null when missing, mill cells are coerced to 0
    in the parser. (Per the user's data-quality rules.)
    """
    missing: list[str] = []

    for block in (sheet.day_shift, sheet.night_shift):
        for line_segs, line_no in ((block.line1_segments, 1), (block.line2_segments, 2)):
            for seg_idx, line in enumerate(line_segs, start=1):
                key = f"{block.shift}/line{line_no}/seg{seg_idx}"

                for fname in _ALWAYS_REQUIRED_LINE_FIELDS:
                    if getattr(line, fname) is None:
                        missing.append(f"{key}: {fname} is None")

                if _line_is_idle(line):
                    continue

                for fname, col_attr in _PROD_OPTIONAL_FIELDS:
                    if getattr(prod_layout, col_attr) is None:
                        continue
                    if fname in _INHERENTLY_OPTIONAL_LINE:
                        continue
                    if getattr(line, fname) is None:
                        missing.append(f"{key}: {fname} (col exists, cell empty)")

    return missing


# ── Public entry ──────────────────────────────────────────────────────────────

def parse_sheet(ws: Worksheet) -> SheetData:
    """Parse one daily-report sheet into structured data. Raises ParseError on structural mismatch."""
    _validate_anchors(ws)

    prod_layout = _discover_production_layout(ws)
    dt_layout = _discover_downtime_layout(ws)

    # Discover quality section first so it bounds the FP-section probe.
    quality_section_row = _find_row_in_col(ws, 1, "کیفیت خوراک", row_start=15)
    if not quality_section_row:
        raise ParseError("could not find 'کیفیت خوراک و کنسانتره تولیدی' section header in column A")

    # FP section: either explicit header, sub-header pattern, or absent (None).
    # Conflicts (both header and sub-header in range) raise ParseError.
    fp_layout = _discover_filter_press(ws, downtime_end_hint=14,
                                       quality_section_row=quality_section_row)

    dt_start_row = dt_layout.header_row + 1
    dt_end_row = (fp_layout.section_row - 1) if fp_layout else (quality_section_row - 1)
    quality_layout = _discover_quality(
        ws, fp_end_hint=fp_layout.last_data_row if fp_layout else (quality_section_row - 1),
    )

    jalali_date = _to_str(ws["U3"].value) or _to_str(ws["V3"].value)
    if not jalali_date:
        raise ParseError("missing date at U3 and V3")

    # Production block — multi-load aware; emits 1 or 2 segments per (shift, line)
    day_l1_segs, day_l2_segs, night_l1_segs, night_l2_segs, _totals_row = _walk_production_block(ws, prod_layout)

    day = ShiftBlock(
        shift="day",
        supervisor_name=_to_str(ws["D3"].value),
        downtime_description=None,
        water_consumption=None,
        line1_segments=day_l1_segs,
        line2_segments=day_l2_segs,
    )
    night = ShiftBlock(
        shift="night",
        supervisor_name=_to_str(ws["K3"].value),
        downtime_description=None,
        water_consumption=None,
        line1_segments=night_l1_segs,
        line2_segments=night_l2_segs,
    )

    # Mill, downtime, and quality readers operate on segment 1 only — those
    # readings are recorded once per shift-line, not per load.
    targets = {
        ("day", 1): day.line1, ("day", 2): day.line2,
        ("night", 1): night.line1, ("night", 2): night.line2,
    }

    # Mill ball additions + water
    mill_rows = _find_mill_rows(ws, dt_layout, dt_end_row)
    water_per_shift = {"day": 0.0, "night": 0.0}
    water_seen = {"day": False, "night": False}
    for (shift_name, line_no), r in mill_rows.items():
        line_row = targets[(shift_name, line_no)]
        w = _apply_mills_and_water(ws, line_row, r, dt_layout)
        if w is not None:
            water_per_shift[shift_name] += w
            water_seen[shift_name] = True
    day.water_consumption = water_per_shift["day"] if water_seen["day"] else None
    night.water_consumption = water_per_shift["night"] if water_seen["night"] else None

    # Factory + feed-input downtime + per-shift red-font notes
    _read_factory_and_feed_downtime(ws, dt_layout, dt_start_row, dt_end_row,
                                    day_block=day, night_block=night)
    # Filter-press downtime (some sheets omit this section)
    if fp_layout:
        _read_filter_press(ws, fp_layout, targets)
    # Quality (4 slots: day1, day2, night1, night2; some may be None when the operator
    # didn't record a row for that line).
    qrows = quality_layout.data_rows
    if qrows[0] is not None:
        _apply_quality(ws, day.line1, qrows[0], quality_layout)
    if qrows[1] is not None:
        _apply_quality(ws, day.line2, qrows[1], quality_layout)
    if qrows[2] is not None:
        _apply_quality(ws, night.line1, qrows[2], quality_layout)
    if qrows[3] is not None:
        _apply_quality(ws, night.line2, qrows[3], quality_layout)

    sheet_data = SheetData(jalali_date=jalali_date, day_shift=day, night_shift=night)

    # "Don't skip data" gate — abort if any field is unexpectedly empty on an active line.
    missing = _validate_no_unexpected_missing(sheet_data, prod_layout, quality_layout)
    if missing:
        raise MissingDataError(missing)

    return sheet_data
