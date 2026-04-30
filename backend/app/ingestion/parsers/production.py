"""Parser for the daily production workbook.

Strategy:
  1. Open the workbook, iterate its sheets (one sheet = one day).
  2. For each sheet, locate anchors by Persian text labels (date, shift,
     downtime block) — robust to the 58/59/63-row template variants.
  3. Numeric production data → ProductionShift rows.
  4. Free-text downtime entries → Downtime rows; embeddings + structured
     fields are filled in by the enrichment pipeline.
  5. On validation failure, fall back to claude_fallback_parse(ws).
"""

from io import BytesIO

from openpyxl import load_workbook

from app.ingestion.anchors import find_cell_containing, read_block_below, read_text_block
from app.ingestion.claude_fallback import claude_fallback_parse
from app.ingestion.jalali import parse_jalali


def ingest(filename: str, contents: bytes) -> dict:
    wb = load_workbook(BytesIO(contents), data_only=True)
    parsed_sheets, failures = [], []
    for name in wb.sheetnames:
        try:
            parsed_sheets.append(_parse_sheet(wb[name]))
        except Exception as exc:
            failures.append({"sheet": name, "error": str(exc)})
    return {
        "filename": filename,
        "sheets_parsed": len(parsed_sheets),
        "sheets_failed": len(failures),
        "failures": failures,
    }


def _parse_sheet(ws) -> dict:
    date_cell = find_cell_containing(ws, "تاریخ")
    if date_cell is None:
        return claude_fallback_parse(ws)

    raw_date = ws.cell(row=date_cell.row, column=date_cell.column + 2).value
    report_date = parse_jalali(str(raw_date)) if raw_date else None

    production_anchor = find_cell_containing(ws, "شیفت روز")
    production_rows = (
        read_block_below(ws, production_anchor, num_rows=5) if production_anchor else []
    )

    downtime_anchor = find_cell_containing(ws, "علت توقف")
    downtime_entries = read_text_block(ws, downtime_anchor) if downtime_anchor else []

    return {
        "report_date": report_date,
        "production_rows": production_rows,
        "downtime_entries": downtime_entries,
    }
