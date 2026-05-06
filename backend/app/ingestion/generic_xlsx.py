"""
Generic xlsx dumper — captures every non-empty cell from every sheet of a workbook.

This is the fallback (and complement) to the structured production parser. The
typed parser writes named columns into production_shift / downtime; this
dumper walks every worksheet and emits one record per cell so every value is
preserved in queryable SQL form (raw_xlsx_cells).

Records have value coerced into one of three columns:
  - value_num   for numbers
  - value_date  for Excel dates
  - value_text  fallback (always populated with the str() form too — that way
                 the agent can do ILIKE searches uniformly)
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.utils import get_column_letter


def _coerce(value):
    """
    Map an openpyxl cell value to (value_text, value_num, value_date).
    Always populate value_text with a string form so substring queries work.
    """
    if value is None:
        return None, None, None

    # bool is a subclass of int in Python — handle it before the int branch.
    if isinstance(value, bool):
        return ("true" if value else "false"), (1.0 if value else 0.0), None

    if isinstance(value, datetime):
        # Excel datetimes get split: store the date part for queries, full text
        # for display. value_num holds the Excel serial (days since 1899-12-30).
        return value.isoformat(sep=" "), None, value.date()

    if isinstance(value, date):
        return value.isoformat(), None, value

    if isinstance(value, (int, float)):
        return str(value), float(value), None

    s = str(value).strip()
    if not s:
        return None, None, None

    # Try to parse Persian/Arabic digits as well — some sheets mix them in.
    num = _maybe_number(s)
    return s, num, None


_PERSIAN_DIGIT_TRANS = str.maketrans(
    "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
    "01234567890123456789",
)


def _maybe_number(s: str) -> float | None:
    """Return float(s) if the string is numeric (also handling Persian digits, %, commas), else None."""
    if not s:
        return None
    cleaned = s.translate(_PERSIAN_DIGIT_TRANS).replace(",", "").replace(" ", "")
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
    if cleaned in ("", "-", "—", "#DIV/0!", "#N/A", "#REF!", "#VALUE!", "#NAME?", "#NULL!"):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def iter_cells(path: str | Path) -> Iterator[dict]:
    """
    Yield one dict per non-empty cell across every sheet of the workbook.

    Yielded dict matches the RawXlsxCell columns (sans file_id, which the
    orchestrator fills in).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text, num, dt = _coerce(cell.value)
                    if text is None and num is None and dt is None:
                        continue
                    addr = f"{get_column_letter(cell.column)}{cell.row}"
                    yield {
                        "sheet_name": sheet_name[:120],
                        "sheet_index": sheet_index,
                        "row_idx": cell.row,
                        "col_idx": cell.column,
                        "cell_address": addr,
                        "value_text": text,
                        "value_num": num,
                        "value_date": dt,
                        "is_formula": False,  # data_only=True → we already get evaluated values
                    }
    finally:
        wb.close()
