"""
Ingest a daily-production-report workbook (1401-1405 family) into the database.

Each sheet is processed inside its own savepoint so a single bad sheet does
not block the rest of the file. Errors are reported with the *phase* in which
they occurred:
    'load'        — workbook could not be opened
    'parse'       — sheet structure didn't match the template
    'date'        — Jalali date couldn't be converted to Gregorian
    'day_shift'   — error writing the day-shift block
    'night_shift' — error writing the night-shift block

Idempotency:
  - supervisors / loads:        lookup-or-create by name / code
  - shifts:                     UNIQUE(date, shift) — re-used if present
  - line_shift_reports:         UNIQUE(shift_id, line_number) — updated in place
  - downtime children (factory / input_feed / filter_press): cleared and
    re-inserted on every ingest of a (shift, line). The corresponding
    line_shift_report is the parent.

`ingest_sheet()` returns a `SheetResult` that, on success, includes the IDs of
the two shifts and the four line_shift_reports written for that sheet.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.enrichment.downtimes import enrich
from app.ingestion.jalali import to_gregorian
from app.ingestion.loads_normalize import normalize_load_code
from app.ingestion.parser_v1 import (
    LineDowntimes,
    LineRow,
    MissingDataError,
    ParseError,
    SheetData,
    ShiftBlock,
    parse_sheet,
)
from app.models import (
    FactoryDowntime,
    FilterPressDowntime,
    InputFeedDowntime,
    LineShiftReport,
    Load,
    Shift,
    Supervisor,
)


# ── Result types ──────────────────────────────────────────────────────────────

@dataclass
class SheetIds:
    day_shift_id: int
    day_line1_lsr_id: int
    day_line2_lsr_id: int
    night_shift_id: int
    night_line1_lsr_id: int
    night_line2_lsr_id: int


@dataclass
class SheetResult:
    sheet_name: str
    status: str                                  # 'ok' | 'error'
    phase: Optional[str] = None
    error: Optional[str] = None
    jalali_date: Optional[str] = None
    rows_written: int = 0                        # 0..4 line_shift_reports
    ids: Optional[SheetIds] = None


@dataclass
class FileResult:
    file: str
    sheets: list[SheetResult] = field(default_factory=list)

    @property
    def ok(self) -> int:
        return sum(1 for s in self.sheets if s.status == "ok")

    @property
    def failed(self) -> int:
        return sum(1 for s in self.sheets if s.status == "error")


# ── Lookup-or-create helpers ──────────────────────────────────────────────────

def _get_or_create_supervisor(db: Session, name: Optional[str]) -> Optional[Supervisor]:
    if not name:
        return None
    row = db.execute(select(Supervisor).where(Supervisor.name == name)).scalar_one_or_none()
    if row:
        return row
    row = Supervisor(name=name)
    db.add(row)
    db.flush()
    return row


def _get_or_create_load(db: Session, code: Optional[str]) -> Optional[Load]:
    if not code:
        return None
    canonical = normalize_load_code(code)
    if not canonical:
        return None
    row = db.execute(select(Load).where(Load.code == canonical)).scalar_one_or_none()
    if row:
        return row
    row = Load(code=canonical)
    db.add(row)
    db.flush()
    return row


def _get_or_create_shift(db: Session, gregorian_date, jalali_date: str, shift_name: str,
                         supervisor: Optional[Supervisor], water_consumption: Optional[float],
                         downtime_description: Optional[str]) -> Shift:
    row = db.execute(
        select(Shift).where(Shift.date == gregorian_date, Shift.shift == shift_name)
    ).scalar_one_or_none()
    if row:
        if supervisor and row.supervisor_id is None:
            row.supervisor_id = supervisor.id
        if water_consumption is not None:
            row.water_consumption = water_consumption
        if downtime_description is not None:
            row.downtime_description = downtime_description
        return row
    row = Shift(
        date=gregorian_date,
        jalali_date=jalali_date,
        shift=shift_name,
        supervisor_id=supervisor.id if supervisor else None,
        water_consumption=water_consumption,
        downtime_description=downtime_description,
    )
    db.add(row)
    db.flush()
    return row


# ── Line shift report upsert ─────────────────────────────────────────────────

def _upsert_report(db: Session, shift: Shift, line: LineRow, load: Optional[Load], load_segment: int) -> LineShiftReport:
    """Insert or update the line_shift_report for (shift, line, load_segment). Returns the row."""
    existing = db.execute(
        select(LineShiftReport).where(
            LineShiftReport.shift_id == shift.id,
            LineShiftReport.line_number == line.line_number,
            LineShiftReport.load_segment == load_segment,
        )
    ).scalar_one_or_none()

    fields = dict(
        load_id=load.id if load else None,
        # Production
        input_feed_tonnage=line.input_feed_tonnage,
        production_tonnage=line.production_tonnage,
        recovery=line.recovery,
        operation_hour=line.operation_hour,
        downtime_hour=line.downtime_hour,
        ton_per_hour=line.ton_per_hour,
        drum_filter_1_hour=line.drum_filter_1_hour,
        drum_filter_2_hour=line.drum_filter_2_hour,
        filter_press_operation_hour=line.filter_press_operation_hour,
        filter_press_downtime_hour=line.filter_press_downtime_hour,
        flocculant_consumption_grams=line.flocculant_consumption_grams,
        flocculant_type=line.flocculant_type,
        # Mills
        primary_mill_30=line.primary_mill_30,
        primary_mill_40=line.primary_mill_40,
        primary_mill_50=line.primary_mill_50,
        primary_mill_60=line.primary_mill_60,
        secondary_mill_25=line.secondary_mill_25,
        secondary_mill_30=line.secondary_mill_30,
        secondary_mill_40=line.secondary_mill_40,
        secondary_mill_50=line.secondary_mill_50,
        # Quality — Fe / Feo
        fe_input_feed=line.fe_input_feed,
        feo_input_feed=line.feo_input_feed,
        fe_concentrate=line.fe_concentrate,
        feo_concentrate=line.feo_concentrate,
        fe_thickener_tailing=line.fe_thickener_tailing,
        feo_thickener_tailing=line.feo_thickener_tailing,
        fe_first_ballmill_output=line.fe_first_ballmill_output,
        feo_first_ballmill_output=line.feo_first_ballmill_output,
        # K80
        k80_size_input_feed=line.k80_size_input_feed,
        k80_size_primary_ballmill=line.k80_size_primary_ballmill,
        k80_size_secondary_ballmill=line.k80_size_secondary_ballmill,
        k80_size_hydrocyclone_overflow_1=line.k80_size_hydrocyclone_overflow_1,
        k80_size_hydrocyclone_overflow_2=line.k80_size_hydrocyclone_overflow_2,
        k80_size_tailing=line.k80_size_tailing,
        k80_size_concentrate=line.k80_size_concentrate,
        # Recoveries / moisture
        dry_weight_recovery=line.dry_weight_recovery,
        metallurgical_recovery=line.metallurgical_recovery,
        separation_efficiency=line.separation_efficiency,
        input_feed_moisture=line.input_feed_moisture,
        concentrate_moisture=line.concentrate_moisture,
        filter_press_cake_moisture=line.filter_press_cake_moisture,
    )

    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        db.flush()
        return existing
    new = LineShiftReport(
        shift_id=shift.id,
        line_number=line.line_number,
        load_segment=load_segment,
        **fields,
    )
    db.add(new)
    db.flush()
    return new


_DOWNTIME_LOG = logging.getLogger(__name__)


def _safe_enrich(description: Optional[str]) -> dict[str, Any]:
    """Wrap enrich() so ingest never fails when enrichment hits a snag.

    The fallback returns valid values for NOT-NULL-bound columns so the
    constraints added in migration 003 can't be violated.
    """
    try:
        return enrich(description)
    except Exception as exc:
        _DOWNTIME_LOG.warning("downtime enrichment failed (%s): %s",
                              (description or "")[:60], exc)
        return {
            "embedding": None,
            "category": "other",
            "department_tag": None,
            "equipment_codes": None,
            "start_time": None,
            "end_time": None,
            "is_planned": False,
        }


def _replace_downtimes(db: Session, lsr: LineShiftReport, downs: LineDowntimes) -> None:
    """Wipe and re-insert the three downtime child tables for this LSR.

    Each inserted row is enriched in-place via app.enrichment.downtimes.enrich
    (embedding + category + extracted fields). Enrichment failures degrade
    gracefully — they never block ingest.
    """
    db.execute(delete(InputFeedDowntime).where(InputFeedDowntime.line_shift_report_id == lsr.id))
    db.execute(delete(FilterPressDowntime).where(FilterPressDowntime.line_shift_report_id == lsr.id))
    db.execute(delete(FactoryDowntime).where(FactoryDowntime.line_shift_report_id == lsr.id))
    db.flush()

    factory_rows: list[FactoryDowntime] = []
    for ev in downs.factory:
        e = _safe_enrich(ev.description)
        row = FactoryDowntime(
            line_shift_report_id=lsr.id,
            description=ev.description,
            duration=ev.duration_minutes,
            **e,
        )
        db.add(row)
        factory_rows.append(row)
    if factory_rows:
        db.flush()                                    # populate IDs

    for ev, link_idx in downs.input_feed:
        link_id = factory_rows[link_idx].id if (link_idx is not None and link_idx < len(factory_rows)) else None
        e = _safe_enrich(ev.description)
        db.add(InputFeedDowntime(
            line_shift_report_id=lsr.id,
            factory_downtime_id=link_id,
            description=ev.description,
            duration=ev.duration_minutes,
            **e,
        ))

    for ev in downs.filter_press:
        e = _safe_enrich(ev.description)
        db.add(FilterPressDowntime(
            line_shift_report_id=lsr.id,
            description=ev.description,
            duration=ev.duration_minutes,
            **e,
        ))


# ── Per-shift driver ─────────────────────────────────────────────────────────

def _ingest_shift_block(db: Session, gregorian_date, jalali_date: str,
                        block: ShiftBlock) -> tuple[Shift, LineShiftReport, LineShiftReport]:
    """Insert one Shift and one LineShiftReport per (line, segment).
    Returns (shift, segment-1 line1 LSR, segment-1 line2 LSR) for the test driver."""
    supervisor = _get_or_create_supervisor(db, block.supervisor_name)
    shift = _get_or_create_shift(db, gregorian_date, jalali_date, block.shift, supervisor,
                                 water_consumption=block.water_consumption,
                                 downtime_description=block.downtime_description)

    seg1_lsrs: dict[int, LineShiftReport] = {}
    for line_no, segs in ((1, block.line1_segments), (2, block.line2_segments)):
        for seg_idx, line in enumerate(segs, start=1):
            load = _get_or_create_load(db, line.load_code)
            lsr = _upsert_report(db, shift, line, load, load_segment=seg_idx)
            _replace_downtimes(db, lsr, line.downtimes)
            if seg_idx == 1:
                seg1_lsrs[line_no] = lsr

    return shift, seg1_lsrs[1], seg1_lsrs[2]


# ── Per-sheet driver ─────────────────────────────────────────────────────────

def ingest_sheet(db: Session, ws, sheet_name: str) -> SheetResult:
    try:
        data = parse_sheet(ws)
    except MissingDataError as e:
        return SheetResult(sheet_name=sheet_name, status="error", phase="missing_data",
                           error=" || ".join(e.missing))
    except ParseError as e:
        return SheetResult(sheet_name=sheet_name, status="error", phase="parse", error=str(e))
    except Exception as e:                                  # pragma: no cover
        return SheetResult(sheet_name=sheet_name, status="error", phase="parse",
                           error=f"{type(e).__name__}: {e}")

    gregorian = to_gregorian(data.jalali_date)
    if gregorian is None:
        return SheetResult(sheet_name=sheet_name, status="error", phase="date",
                           jalali_date=data.jalali_date,
                           error=f"could not convert jalali date {data.jalali_date!r}")

    sp = db.begin_nested()
    try:
        try:
            day_shift, day_lsr1, day_lsr2 = _ingest_shift_block(db, gregorian, data.jalali_date, data.day_shift)
        except Exception as e:
            sp.rollback()
            return SheetResult(sheet_name=sheet_name, status="error", phase="day_shift",
                               jalali_date=data.jalali_date, error=f"{type(e).__name__}: {e}")
        try:
            night_shift, night_lsr1, night_lsr2 = _ingest_shift_block(db, gregorian, data.jalali_date, data.night_shift)
        except Exception as e:
            sp.rollback()
            return SheetResult(sheet_name=sheet_name, status="error", phase="night_shift",
                               jalali_date=data.jalali_date, error=f"{type(e).__name__}: {e}")
        sp.commit()
    except Exception:                                       # pragma: no cover
        sp.rollback()
        raise

    return SheetResult(
        sheet_name=sheet_name, status="ok",
        jalali_date=data.jalali_date, rows_written=4,
        ids=SheetIds(
            day_shift_id=day_shift.id,
            day_line1_lsr_id=day_lsr1.id,
            day_line2_lsr_id=day_lsr2.id,
            night_shift_id=night_shift.id,
            night_line1_lsr_id=night_lsr1.id,
            night_line2_lsr_id=night_lsr2.id,
        ),
    )


# ── Per-file driver ───────────────────────────────────────────────────────────

def ingest_workbook(db: Session, path: str | Path,
                    sheet_names: Optional[list[str]] = None) -> FileResult:
    p = Path(path)
    result = FileResult(file=str(p))

    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=False)
    except Exception as e:
        result.sheets.append(SheetResult(sheet_name="<workbook>", status="error",
                                         phase="load", error=f"{type(e).__name__}: {e}"))
        return result

    titles = sheet_names if sheet_names is not None else wb.sheetnames
    for title in titles:
        if title not in wb.sheetnames:
            result.sheets.append(SheetResult(sheet_name=title, status="error", phase="load",
                                             error="sheet not found in workbook"))
            continue
        ws = wb[title]
        sr = ingest_sheet(db, ws, title)
        result.sheets.append(sr)
        if sr.status == "ok":
            db.commit()
    return result
