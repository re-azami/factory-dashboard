"""
Parser for factory daily production Excel workbooks.

Each sheet is one day's report (date in sheet name, e.g. "01_05" = 1405/01/05).

Returns:
  - daily_report:    one dict per sheet (date + header info)
  - production_rows: one dict per (shift, line) — wide row with ALL metrics
  - downtime_rows:   one dict per stop event (factory / feed_input / filter_press)
  - raw_cells:       one dict per sheet with every non-empty cell as JSON

Strategy: anchor-based parsing. We find Persian label cells and read data
relative to them. This is robust to template version changes.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.jalali import to_gregorian


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_num(value: Any) -> float | None:
    """Return a float if cell holds a number, else None. Treats Excel errors as None."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "-", "—", "#DIV/0!", "#N/A", "#REF!", "#VALUE!", "#NAME?", "#NULL!"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _clean_str(value: Any) -> str | None:
    """Return cleaned string or None."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "-", "—", "#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return None
    return s


def _normalize(text: str) -> str:
    """Collapse all whitespace (including newlines) so anchor matching is forgiving."""
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _find_cell(ws: Worksheet, label: str):
    """Find the first cell whose value contains `label` (whitespace-insensitive)."""
    target = _normalize(label)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and target in _normalize(cell.value):
                return cell
    return None


def _find_all_cells(ws: Worksheet, label: str):
    """Find all cells whose value contains `label` (whitespace-insensitive)."""
    target = _normalize(label)
    found = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and target in _normalize(cell.value):
                found.append(cell)
    return found


# ── Date and header extraction ────────────────────────────────────────────────

def _extract_header(ws: Worksheet, sheet_name: str, source_file: str) -> dict | None:
    """Pull date + supervisors + batch code from the sheet header."""
    anchor = _find_cell(ws, "تاریخ")
    jalali_str, gregorian = None, None

    if anchor:
        for offset in (1, 2, 3):
            candidate = ws.cell(row=anchor.row, column=anchor.column + offset)
            val = str(candidate.value or "").strip()
            if re.match(r"\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}", val):
                jalali_str = val
                gregorian = to_gregorian(val)
                break

    # Fallback: parse date from sheet name (e.g. "01_05" → 1405/01/05)
    # Sheet names are MM_DD format in year 1405 (the year of the file).
    if gregorian is None:
        m = re.match(r"(\d{1,2})_(\d{1,2})", sheet_name)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            jalali_str = f"1405/{month:02d}/{day:02d}"
            gregorian = to_gregorian(jalali_str)

    if gregorian is None:
        return None

    # Batch code is usually in the last data row (row 58 in 01_05)
    batch_code = None
    for row in ws.iter_rows():
        for cell in row:
            val = _clean_str(cell.value)
            if val and ("MIX" in val or "MAH" in val) and len(val) > 10:
                batch_code = val
                break
        if batch_code:
            break

    return {
        "report_date": gregorian,
        "jalali_date": jalali_str,
        "sheet_name": sheet_name,
        "source_file": source_file,
        "batch_code": batch_code,
        "supervisors": None,  # could parse row 3 if needed later
    }


# ── Production rows ───────────────────────────────────────────────────────────

# Maps Persian shift label → canonical name
_SHIFT_LABELS = {
    "شیفت روز": "day",
    "شیفت شب": "night",
    "جمع کل": "total",
    "کل": "total",
}

# Production columns (rows 7-11 area). Column letters from Excel inspection.
# Letter → (model field name, type)
_PRODUCTION_COLS = {
    "C": ("daily_feed_tonnage", "num"),
    "E": ("daily_concentrate_tonnage", "num"),
    "G": ("daily_recovery_percent", "num"),
    "H": ("ore_grade_code", "str"),
    "I": ("monthly_feed_tonnage", "num"),
    "K": ("monthly_concentrate_tonnage", "num"),
    "N": ("monthly_recovery_percent", "num"),
    "P": ("yearly_feed_tonnage", "num"),
    "R": ("yearly_concentrate_tonnage", "num"),
    "T": ("yearly_recovery_percent", "num"),
    "U": ("throughput_ton_per_hour", "num"),
    "V": ("factory_operation_hours", "num"),
    "W": ("factory_downtime_hours", "num"),
    "X": ("feed_input_operation_hours", "num"),
    "Y": ("feed_input_downtime_hours", "num"),
    "Z": ("drum_filter_1_hours", "num"),
    "AA": ("drum_filter_2_hours", "num"),
    "AB": ("filter_press_operation_hours", "num"),
    "AC": ("filter_press_downtime_hours", "num"),
    "AD": ("flocculant_grams", "num"),
    "AE": ("flocculant_type", "str"),
}

# Quality rows (rows 50-57 area). Same structure: shift+line in cols A-B.
_QUALITY_COLS = {
    "C": ("feed_fe_percent", "num"),
    "D": ("feed_feo_percent", "num"),
    "E": ("concentrate_fe_percent", "num"),
    "G": ("concentrate_feo_percent", "num"),
    "I": ("tailings_fe_percent", "num"),
    "J": ("tailings_feo_percent", "num"),
    "K": ("feed_k80_microns", "num"),
    "M": ("primary_mill_output", "num"),
    "O": ("secondary_mill_output", "num"),
    "P": ("hydrocyclone_1_overflow", "num"),
    "Q": ("hydrocyclone_2_overflow", "num"),
    "R": ("tailings_k80_microns", "num"),
    "T": ("concentrate_k80_microns", "num"),
    "U": ("dry_weight_recovery_percent", "num"),
    "X": ("assay_recovery_percent", "num"),
    "Y": ("separation_efficiency_percent", "num"),
    "Z": ("feed_moisture_percent", "num"),
    "AA": ("concentrate_moisture_percent", "num"),
    "AB": ("filter_cake_moisture_percent", "num"),
    "AC": ("primary_mill_output_fe_percent", "num"),
    "AD": ("primary_mill_output_feo_percent", "num"),
}


def _col_letter_to_index(letter: str) -> int:
    """A → 1, B → 2, ..., AA → 27, AB → 28, ..."""
    n = 0
    for c in letter:
        n = n * 26 + (ord(c.upper()) - ord("A") + 1)
    return n


def _parse_shift_line(shift_val: str, line_val: str) -> tuple[str | None, int | None]:
    """Map raw cell text to canonical (shift, line)."""
    shift = None
    for label, canonical in _SHIFT_LABELS.items():
        if label in shift_val:
            shift = canonical
            break

    line = None
    line_match = re.search(r"(\d+)", line_val)
    if line_match:
        line = int(line_match.group(1))

    return shift, line


def _read_block(ws: Worksheet, start_row: int, end_row: int, col_map: dict, jalali_str: str, gregorian, source_file: str) -> dict:
    """
    Read a block of rows (each row = one shift+line) and return a dict keyed by (shift, line).
    Used for both production block and quality block.
    """
    out = {}
    current_shift = None  # shift label may span multiple lines via merged cell

    for row in range(start_row, end_row + 1):
        shift_val = str(ws.cell(row=row, column=1).value or "").strip()
        line_val = str(ws.cell(row=row, column=2).value or "").strip()

        # Update current shift if this row has one (sticky for merged cells)
        new_shift, line = _parse_shift_line(shift_val, line_val)
        if new_shift:
            current_shift = new_shift

        if current_shift is None or line is None:
            # Try total row (no line specified)
            if "جمع" in shift_val or "کل" in shift_val:
                key = ("total", None)
            else:
                continue
        else:
            key = (current_shift, line)

        if key not in out:
            out[key] = {
                "report_date": gregorian,
                "jalali_date": jalali_str,
                "shift": key[0],
                "line": key[1],
                "source_file": source_file,
            }

        # Extract each mapped column
        for col_letter, (field_name, kind) in col_map.items():
            cell_value = ws.cell(row=row, column=_col_letter_to_index(col_letter)).value
            if kind == "num":
                v = _clean_num(cell_value)
                if v is not None:
                    out[key][field_name] = v
            else:
                v = _clean_str(cell_value)
                if v is not None:
                    out[key][field_name] = v

    return out


def _extract_production_and_quality(ws: Worksheet, jalali_str: str, gregorian, source_file: str) -> list[dict]:
    """Extract production block (rows 7-11) and quality block (rows 50-57), merge by (shift, line)."""
    # Find row anchors for the two blocks
    day_anchor = _find_cell(ws, "شیفت روز")
    if day_anchor is None:
        return []

    production_start = day_anchor.row
    production_end = production_start + 6  # ~5-6 rows of data

    # Quality section starts after the downtime sections
    quality_anchor = _find_cell(ws, "کیفیت خوراک")
    quality_start = quality_anchor.row + 2 if quality_anchor else None
    quality_end = quality_start + 8 if quality_start else None

    production_data = _read_block(ws, production_start, production_end, _PRODUCTION_COLS, jalali_str, gregorian, source_file)

    if quality_start:
        quality_data = _read_block(ws, quality_start, quality_end, _QUALITY_COLS, jalali_str, gregorian, source_file)
        # Merge quality fields into production rows
        for key, q_row in quality_data.items():
            if key in production_data:
                production_data[key].update({k: v for k, v in q_row.items() if k not in ("report_date", "jalali_date", "shift", "line", "source_file")})
            else:
                production_data[key] = q_row

    return list(production_data.values())


# ── Downtime extraction ───────────────────────────────────────────────────────

# Each section has its own anchor text + column layout.
# - anchor: how to find the section in the sheet (whitespace-tolerant)
# - text_col_offset: column for the Persian text, relative to the anchor column
# - duration_col_offset: column for the numeric duration, relative to the anchor column
# - row_span: how many rows to scan below the anchor row
# - shift_col: which column holds 'شیفت روز' / 'شیفت شب' for this section (None = none)
_DOWNTIME_SECTIONS = [
    # Factory downtime: anchor in column L (e.g. "دلایل توقف ختکا" at L13).
    # Text in column N, duration in column W. Shift label in column L.
    {
        "section": "factory",
        "anchors": ["دلایل توقف ختک", "دلایل توقف کارخانه", "علت توقف کارخانه"],
        "text_col_offset": 2,
        "duration_col_offset": 11,
        "row_span": 25,
        "shift_col_offset": 0,   # shift in same column as anchor (L)
        "line_col_offset": 1,    # line in M
    },
    # Feed input downtime: anchor in column X (e.g. "علت توقف فید ورودی" at X14).
    # Text in column X (same as anchor), duration in column AD.
    {
        "section": "feed_input",
        "anchors": ["علت توقف فید", "علت توقف  فید"],
        "text_col_offset": 0,
        "duration_col_offset": 6,
        "row_span": 25,
        "shift_col_offset": -23,  # shift back at column A
        "line_col_offset": -22,
    },
    # Filter press downtime appears twice per sheet: day (anchor at A41) and night (anchor at Q41).
    # Day: text in C, duration in K.    Offsets from A: text +2, dur +10.
    # Night: text in T, duration in AD. Offsets from Q: text +3, dur +13.
    {
        "section": "filter_press",
        "anchors": ["دلایل توقف فیلتر پرس"],
        "text_col_offset": 2,    # works for the day-shift version (A→C)
        "duration_col_offset": 10,
        "row_span": 8,
        "shift_col_offset": None,  # filter press doesn't repeat shift in column
        "line_col_offset": None,
        # For night-shift anchor at Q, the text is +3 not +2 — handled below by scanning multiple offsets.
        "alt_text_offsets": [2, 3],
        "alt_duration_offsets": [10, 13],
    },
]


def _extract_downtime(ws: Worksheet, jalali_str: str, gregorian, source_file: str) -> list[dict]:
    """
    Walk every downtime section and extract real stop events.
    A row counts as an event only if it has Persian text >= 10 chars AND a positive duration.
    """
    rows = []
    claimed = set()  # (row, text_col) tuples to prevent double-add

    for cfg in _DOWNTIME_SECTIONS:
        text_offsets = cfg.get("alt_text_offsets", [cfg["text_col_offset"]])
        dur_offsets = cfg.get("alt_duration_offsets", [cfg["duration_col_offset"]])

        for anchor_text in cfg["anchors"]:
            for anchor in _find_all_cells(ws, anchor_text):
                # Try each (text_offset, duration_offset) pair — handles A vs Q anchor variants
                for text_off, dur_off in zip(text_offsets, dur_offsets):
                    text_col = anchor.column + text_off
                    dur_col = anchor.column + dur_off

                    if text_col < 1 or dur_col < 1:
                        continue

                    for r_off in range(1, cfg["row_span"] + 1):
                        row_num = anchor.row + r_off
                        if row_num > ws.max_row:
                            break
                        if (row_num, text_col) in claimed:
                            continue

                        text_val = _clean_str(ws.cell(row=row_num, column=text_col).value)
                        if not text_val or len(text_val) < 10:
                            continue
                        if not re.search(r"[؀-ۿ]", text_val):
                            continue
                        # Skip header rows ("علت توقف", "مدت توقف")
                        norm = _normalize(text_val)
                        if "علت توقف" in norm or "مدت توقف" in norm or "دلایل توقف" in norm:
                            continue

                        duration_val = _clean_num(ws.cell(row=row_num, column=dur_col).value)
                        if duration_val is None or duration_val <= 0 or duration_val > 10000:
                            continue

                        # Shift label (sticky merged cell — scan up to find it)
                        shift = None
                        line = None
                        if cfg.get("shift_col_offset") is not None:
                            sh_col = anchor.column + cfg["shift_col_offset"]
                            if sh_col >= 1:
                                # walk up from this row to find the most recent shift label
                                for back in range(0, 20):
                                    sh_val = _normalize(ws.cell(row=row_num - back, column=sh_col).value or "")
                                    if "روز" in sh_val:
                                        shift = "day"
                                        break
                                    if "شب" in sh_val:
                                        shift = "night"
                                        break
                        if cfg.get("line_col_offset") is not None:
                            ln_col = anchor.column + cfg["line_col_offset"]
                            if ln_col >= 1:
                                for back in range(0, 20):
                                    ln_val = _normalize(ws.cell(row=row_num - back, column=ln_col).value or "")
                                    m = re.search(r"خط\s*(\d)", ln_val)
                                    if m:
                                        line = int(m.group(1))
                                        break

                        # For filter press, infer shift from anchor column position
                        if cfg["section"] == "filter_press" and shift is None:
                            shift = "day" if anchor.column <= 5 else "night"

                        # Extract start/end time from the text if present (HH:MM)
                        time_match = re.search(r"(\d{1,2}:\d{2}).*?(\d{1,2}:\d{2})", text_val)
                        start_time = time_match.group(1) if time_match else None
                        end_time = time_match.group(2) if time_match else None

                        rows.append({
                            "report_date": gregorian,
                            "jalali_date": jalali_str,
                            "section": cfg["section"],
                            "shift": shift,
                            "line": line,
                            "raw_text": text_val,
                            "duration_minutes": int(duration_val),
                            "equipment_code": None,
                            "fault_category": None,
                            "start_time": start_time,
                            "end_time": end_time,
                            "source_file": source_file,
                        })
                        claimed.add((row_num, text_col))

    return rows


# ── Raw cell dump ─────────────────────────────────────────────────────────────

def _dump_cells(ws: Worksheet) -> dict:
    """Dump every non-empty cell as {coordinate: value} for the JSONB backup."""
    out = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                value = cell.value
                # Convert non-JSON-friendly types to strings
                if not isinstance(value, (str, int, float, bool)):
                    value = str(value)
                out[cell.coordinate] = value
    return out


# ── Public API ────────────────────────────────────────────────────────────────

@dataclass
class ParseResult:
    source_file: str
    sheets_parsed: int = 0
    sheets_failed: int = 0
    daily_reports: list[dict] = field(default_factory=list)
    production_rows: list[dict] = field(default_factory=list)
    downtime_rows: list[dict] = field(default_factory=list)
    raw_cells: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def parse_workbook(path: str | Path) -> ParseResult:
    """Parse all sheets in a production Excel workbook."""
    path = Path(path)
    result = ParseResult(source_file=path.name)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header = _extract_header(ws, sheet_name, path.name)
            if header is None:
                result.errors.append(f"Sheet '{sheet_name}': could not find date")
                result.sheets_failed += 1
                continue

            result.daily_reports.append(header)

            production = _extract_production_and_quality(
                ws, header["jalali_date"], header["report_date"], path.name
            )
            downtime = _extract_downtime(
                ws, header["jalali_date"], header["report_date"], path.name
            )
            cells = _dump_cells(ws)

            result.production_rows.extend(production)
            result.downtime_rows.extend(downtime)
            result.raw_cells.append({
                "report_date": header["report_date"],
                "sheet_name": sheet_name,
                "source_file": path.name,
                "cells": cells,
            })

            result.sheets_parsed += 1

        except Exception as exc:
            result.errors.append(f"Sheet '{sheet_name}': {exc}")
            result.sheets_failed += 1

    wb.close()
    return result
