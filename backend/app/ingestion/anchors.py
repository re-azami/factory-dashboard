"""Anchor-based parsing helpers.

Excel sheets shift rows between template versions but their text labels stay
put. Find cells by content, not by coordinate.
"""

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def find_cell_containing(ws: Worksheet, needle: str) -> Cell | None:
    needle = needle.strip()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and needle in cell.value.strip():
                return cell
    return None


def read_block_below(ws: Worksheet, anchor: Cell, num_rows: int) -> list[list]:
    rows = []
    for r in range(anchor.row + 1, anchor.row + 1 + num_rows):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
    return rows


def read_text_block(ws: Worksheet, anchor: Cell, max_rows: int = 50) -> list[str]:
    """Read non-empty strings starting below an anchor, until a blank gap."""
    out, blanks = [], 0
    for r in range(anchor.row + 1, anchor.row + 1 + max_rows):
        value = ws.cell(row=r, column=anchor.column).value
        if isinstance(value, str) and value.strip():
            out.append(value.strip())
            blanks = 0
        else:
            blanks += 1
            if blanks >= 2:
                break
    return out
