"""Print every non-empty cell of one sheet from each year file (rows 11+)."""
import openpyxl
from pathlib import Path

ROOT = Path("/data/raw/factory")
FILES = ["1405.xlsx", "1404.xlsx", "1403_01_04.xlsx", "1403_04-12.xlsx", "1402.xlsx", "1401.xlsx"]

for fname in FILES:
    p = ROOT / fname
    if not p.exists():
        continue
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"\n========== {fname} | sheet={ws.title!r} | dims={ws.dimensions} ==========")
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=False):
        for c in row:
            if c.value not in (None, ""):
                v = repr(c.value)[:80]
                print(f"  {c.coordinate}: {v}")
