"""Inspect first sheet of each year file to verify cell anchors are stable."""
import openpyxl
from pathlib import Path

ROOT = Path("/data/raw/factory")
FILES = ["1405.xlsx", "1404.xlsx", "1403_01_04.xlsx", "1403_04-12.xlsx", "1402.xlsx", "1401.xlsx"]

for fname in FILES:
    p = ROOT / fname
    if not p.exists():
        print(f"=== {fname}: MISSING ===")
        continue
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        print(f"=== {fname}: LOAD ERROR {e} ===")
        continue
    ws = wb[wb.sheetnames[0]]
    print(f"=== {fname} sheets={len(wb.sheetnames)} first={ws.title!r} dims={ws.dimensions} ===")
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for c in row:
            if c.value not in (None, ""):
                v = repr(c.value)[:60]
                print(f"  {c.coordinate}: {v}")
        print("  --")
