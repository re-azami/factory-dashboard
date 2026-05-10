"""Print all non-empty cells in column A of the failing sheets."""
import openpyxl
from pathlib import Path

CASES = [("1404.xlsx", "10_05"), ("1401.xlsx", "01.04")]
for fname, sheet in CASES:
    p = Path("/data/raw/factory") / fname
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[sheet]
    print(f"\n=== {fname} :: {sheet!r} max_row={ws.max_row} max_col={ws.max_column} ===")
    print("Column A non-empty cells:")
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip():
            print(f"  A{r}: {v!r}")
    print("\nLooking for any cell containing 'فیلتر':")
    for row in ws.iter_rows():
        for c in row:
            if c.value and "فیلتر" in str(c.value):
                print(f"  {c.coordinate}: {c.value!r}")
